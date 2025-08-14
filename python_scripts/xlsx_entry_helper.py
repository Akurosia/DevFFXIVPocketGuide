#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# coding: utf8
from collections import OrderedDict
from io import BytesIO
import os
import json
import logging
import natsort
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from ffxiv_aku import *
from typing import Any
try:
    from python_scripts.constants import DIFFERENT_PRONOUNS, DIFFERENT_PRONOUNSS, LANGUAGES
    from python_scripts.helper import getContentName, seperate_data_into_array, EntryType, _get_coords_relative
    from python_scripts.fileimports import *
except Exception:
    from constants import DIFFERENT_PRONOUNS, DIFFERENT_PRONOUNSS, LANGUAGES
    from helper import getContentName, seperate_data_into_array, EntryType, _get_coords_relative
    from fileimports import *

logger: logging.Logger = logging.getLogger()

def get_header_from_xlsx(local_sheet: Worksheet, local_max_column: int) -> list[str]:
    result: list[str] = []
    for j in range(1, local_max_column + 1):
        result.append(str(local_sheet.cell(row=int(1), column=int(j)).value).replace("None", "").strip())
    return result


def get_data_from_xlsx(local_sheet: Worksheet, local_max_column: int, i: int, elements) -> EntryType:
    entry: EntryType = {} # type: ignore
    # for every column in row add all elements into a dict:
    # local_max_column will ignore last column due to how range is working
    for j in range(1, local_max_column + 1):
        name: str = str(local_sheet.cell(row=int(i), column=int(j)).value).replace("None", "").strip()
        entry[elements[j - 1]] = name
    return entry


def clean_entries_from_single_quotes(entry: EntryType):
    for key, value in entry.items():
        if isinstance(value, str):
            if value.startswith("'"):
                entry[key] = value[1:]
            if value.endswith("'"):
                entry[key] = value[:-1]
    return entry


def make_name_readable(entry: str, x: dict[str, str]) -> str:
    name: str = entry
    name = name.replace("[t]", DIFFERENT_PRONOUNS[str(x["Pronoun"])])
    name = name.replace("[a]", DIFFERENT_PRONOUNSS[str(x["Pronoun"])])
    return name


def workOnQuests(entry: EntryType) -> EntryType:
    quest_id: str = entry["quest_id"]
    if quest_id == "":
        for lang in LANGUAGES:
            entry['quest'][lang] = ""
            entry['quest_location'][lang] = ""
            entry['quest_npc'][lang] = ""
        return entry

    quest: dict[str, str | dict[str, str]] = quests[quest_id]
    for lang in LANGUAGES:
        tmp_quest: str = quests[quest_id][f'Name_{lang}'].replace(" ", "").replace(" ", "") # type: ignore
        #entry[f'quest_{lang}'] = tmp_quest
        entry['quest'][lang] = tmp_quest

    for lang in LANGUAGES:
        npc: str = quest['IssuerStart'][f"Singular_{lang}"] # type: ignore
        if "[a]" in npc or "[t]" in npc:
            npc = make_name_readable(npc, quest['IssuerStart'])
        #entry[f'quest_npc_{lang}'] = npc
        entry['quest_npc'][lang] = npc

    issuer_location: str = quest['IssuerLocation'] # type: ignore
    x, y = _get_coords_relative(issuer_location['X'], issuer_location['Map']['OffsetX'], issuer_location['Z'], issuer_location['Map']['OffsetY'], issuer_location['Map']['SizeFactor'], pixel=False)
    for lang in LANGUAGES:
        try:
            tmp_quest_location = f'{issuer_location['Map']["PlaceName"][f'Name_{lang}']} ({x}, {y})'
            entry['quest_location'][lang] = tmp_quest_location
        except KeyError:
            entry['quest_location'][lang] = ""
            print_color_red(f"[XLSX_ENTRY_HELPER:workOnQuests] Error on loading: {quest['IssuerLocation']=} ({quest_id=})")
    return entry


def getEntriesForRouletts(entry: EntryType) -> EntryType:
    global contentfindercondition
    entry['allianceraid'] = None
    entry['frontier'] = None
    entry['expert'] = None
    entry['guildhest'] = None
    entry['highlevelroulette'] = None
    entry['levelcaproulette'] = None
    entry['leveling'] = None
    entry['main'] = None
    entry['mentor'] = None
    entry['normalraid'] = None
    entry['trial'] = None
    for _, value in contentfindercondition.items():
        if value['Name_de'] == getContentName(entry["title"], "de", entry["difficulty"], entry["instanceType"]):
            #print(value['TerritoryType'])
            entry['type'] = value['ContentType']['Name_de'].lower()
            entry['mapid'] = value['TerritoryType'].get('Name_de', "")
            entry['allianceraid'] = str(value['AllianceRoulette'])
            entry['frontier'] = str(value['FeastTeamRoulette'])
            entry['expert'] = str(value['ExpertRoulette'])
            entry['guildhest'] = str(value['GuildHestRoulette'])
            entry['highlevelroulette'] = str(value['HighLevelRoulette'])
            entry['levelcaproulette'] = str(value['LevelCapRoulette'])
            entry['leveling'] = str(value['LevelingRoulette'])
            entry['main'] = str(value['MSQRoulette'])
            entry['mentor'] = str(value['MentorRoulette'])
            entry['normalraid'] = str(value['NormalRaidRoulette'])
            entry['trial'] = str(value['TrialRoulette'])
            return entry
    return entry


def getBeforeAndAfterContentEntries(orderedContent, entry: EntryType):
    _previous = None
    _next = None
    _type = orderedContent[entry['instanceType']]
    _typeKeys = list(_type)
    for i, k in enumerate(_type):
        if _type[k].endswith(entry['slug_url']):
            if i - 1 >= 0:
                try:
                    _previous = _type[_typeKeys[i - 1]]
                except Exception:
                    pass
            try:
                _next = _type[_typeKeys[i + 1]]
            except Exception:
                pass
            return _previous, _next
    return "", ""


def createNewElements(entry: EntryType) -> EntryType:
    if not entry.get("quest", None):
        entry['quest'] = {}
    if not entry.get("quest_location", None):
        entry['quest_location'] = {}
    if not entry.get("quest_npc", None):
        entry['quest_npc'] = {}
    if not entry.get("titles", None):
        entry['titles'] = {}
    return entry


def getEntryDataOld(sheet: Worksheet, max_column: int, i: int, elements, orderedContent):
    entry: EntryType = get_data_from_xlsx(sheet, max_column, i, elements)
    entry = getEntryData(entry, i, orderedContent)
    return entry


def getEntryData(googledata: EntryType, i: int, orderedContent):
    entry: EntryType = googledata
    entry = clean_entries_from_single_quotes(entry)
    entry = createNewElements(entry)
    entry = workOnQuests(entry)
    entry = getEntriesForRouletts(entry)
    for lang in LANGUAGES:
        tpm_title: str = getContentName(entry["title"], lang, entry["difficulty"], entry["instanceType"])
        #entry[f"title_{lang}"] = tpm_title
        entry["titles"][lang] = tpm_title
    _previous, _next = getBeforeAndAfterContentEntries(orderedContent, entry)
    # remove time from excel datetime
    entry["date"] = str(entry["date"]).replace(" 00:00:00", "").replace("-", ".")
    entry["prev_content"] = _previous
    entry["next_content"] = _next
    entry["line_index"] = i
    seperate_data_into_array("bosse", entry)
    seperate_data_into_array("tags", entry)
    entry['adds'] = []
    entry['mechanics'] = "[]"
    #print(entry)
    return entry


def getPrevAndNextContentOrder(sheet: Worksheet, elements, max_row: int) -> OrderedDict[Any, Any]:
    entry = {}
    for i in range(1, max_row + 1):
        instanceType = str(sheet.cell(row=int(i), column=int(elements.index('instanceType')) + 1).value).replace("None", "")
        if not entry.get(instanceType, None):
            entry[instanceType] = {}
        sortID = str(sheet.cell(row=int(i), column=int(3)).value).replace("None", "")
        addon = str(sheet.cell(row=int(i), column=int(5)).value).replace("None", "")
        #column 7 is the slug_url field while 6 would be the normal slug field
        slug = str(sheet.cell(row=int(i), column=int(7)).value).replace("None", "")
        entry[instanceType][sortID] = "/" + addon + "/" + slug
    return OrderedDict(natsort.natsorted(entry.items()))


def load_workbook_from_url(url: str) -> Workbook:
    file: requests.Response = requests.get(url=url)
    return load_workbook(filename=BytesIO(file.content))


def read_xlsx_file():
    #KEY: str|None = os.environ.get("GDRIVE_APIKEY")
    #if KEY is None:
    #    raise Exception("Missing GDRIVE_APIKEY in Environment")
    #print(KEY)
    #MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #SHEET = "11SBQWYyFnyh19Ku6T1Wr5tNAJvxdze8tvD6m4bzPGIA"
    #r: str = f"https://www.googleapis.com/drive/v3/files/{SHEET}/export?key={KEY}&mimeType={MIME_TYPE}"
    r: str = "https://docs.google.com/spreadsheets/d/e/2PACX-1vT-NhazAByNg-0YtEumVnmJKQjVmHKbuThyAiFLk_3I1asUOCpbto89usgrxaM0Q2-gUqdVYVaKPNNl/pub?output=xlsx"
    wb: Workbook = load_workbook_from_url(r)
    # wb = openpyxl.load_workbook('./guide_ffxiv.xlsx')
    local_sheet: Worksheet = wb['Tabelle1']
    local_max_row: int = local_sheet.max_row
    local_max_column: int = local_sheet.max_column
    return local_sheet, local_max_row, local_max_column


def excel_to_json(sheet: Worksheet) -> dict[str, EntryType]:
    """
    Converts an Excel worksheet to a JSON object.

    Args:
        excel_file (str): Path to the Excel file.
        sheet_name (str, optional): Name of the worksheet to convert. If None, the active sheet is used.
        output_file (str, optional): Path to save the JSON output. If None, the JSON is returned as a string.

    Returns:
        str or None: JSON string or None if saved to file.
    """
    try:
        data: dict[str, EntryType] = {}
        header: list[str] = [str(cell.value) for cell in sheet[1]]  # Get header from the first row

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2): #start at row 2
            row_data: EntryType = {}
            for col_num, cell_value in enumerate(row):
                row_data[header[col_num]] = str(cell_value).replace("None", "").strip()
            data[str(row_num)] = row_data

        return data

    except FileNotFoundError:
        print( f"Error: File '{excel_file}' not found.")
    except KeyError:
        print( f"Error: Sheet '{sheet_name}' not found in the Excel file.")
    except Exception as e:
        print( f"An error occurred: {e}")
    return {}


if __name__ == "__main__":
    #sheet, max_row, max_column = read_xlsx_file()
    #XLSXELEMENTS = get_header_from_xlsx(sheet, max_column)
    entry = {'exclude': '', 'date': '2024.11.12', 'sortid': '7100603144', 'title': 'Jeuno: Die erste Etappe', 'categories': 'dt', 'slug': 'jeuno_die_erste_etappe', 'slug_url': 'jeuno_the_first_walk', 'image': 'ui/icon/112000/112585.tex', 'patchNumber': '7.1', 'patchName': 'Crossroads', 'difficulty': 'Normal', 'CFC_ID': '1015', 'PN_ID': '', 'quest_id': '70769', 'gearset_loot': "['Erzengel']", 'tt_card': "['Schattenlord']", 'orchestrion': "['Vergilbte Notenrolle']", 'orchestrion_material': '', 'mtqvid': '[]', 'mrhvid': '[]', 'hector': '[]', 'mount': '', 'minion': "['Nanolord']", 'instanceType': 'allianzraid', 'mapid': 'z6r1', 'bosse': '["Prishe Von Den Fernen Ketten","Fafnir","Erzengel Gk","Erzengel Hm","Erzengel Mr","Erzengel Tt","Erzengel Ev","Schattenlord"]', 'tags': '[]', 'teamcraftlink': '30144', 'garlandtoolslink': '30144', 'gamerescapelink': 'Jeuno:_The_First_Walk', 'done': '', 'quest': {'de': 'Am Kreuzweg der Welten', 'en': 'An Otherworldly Encounter', 'fr': 'À la croisée des mondes', 'ja': '交わる世界', 'cn': '交错的世界', 'ko': '교차하는 세계'}, 'quest_location': {'de': ''}, 'quest_npc': {'de': 'Hoobigo-Gesandter', 'en': 'Hoobigo messenger', 'fr': 'messager hoobigo', 'ja': '使いのフビゴ族', 'cn': '霍比格族使者', 'ko': '후비고족 심부름꾼'}, 'titles': {}}
    print_pretty_json(workOnQuests(entry))
