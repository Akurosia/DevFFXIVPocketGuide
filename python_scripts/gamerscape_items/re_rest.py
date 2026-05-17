from ffxiv_aku import *
import re
from openpyxl import load_workbook
import os
from io import BytesIO
from bs4 import BeautifulSoup
import requests

enitems = loadDataTheQuickestWay("Item.en.json")
enitem_lookup = {
    value.get('Name', '').lower(): key
    for key, value in enitems.items()
    if value.get('Name', '')
}
enitem_lookup.update({
    value.get('Singular', '').lower(): key
    for key, value in enitems.items()
    if value.get('Singular', '')
})
SKIP_TITLE_PARTS = (
    "#Acquired from Duty",
    "#Sold by Merchant",
    "#Rewarded from Quests",
    "#Rewarded from Treasure Map",
    "#Potentially received from",
    "#Acquired from Airship Voyage",
    "#Logging",
    "#Harvesting",
    "#Mining",
    "#Quarrying",
    "#Rewarded from Levequests",
    "#Acquired from Gardening",
    "#Acquired from Venture",
    "#Dropped by Monsters",
    "#Acquired from Reduction",
    "#Miscellaneous Acquisition",
    "#FATE",
    "#Recipes",
    " (Lost Action)",
    "#Acquired from Desynth",
)
SKIP_TITLES = {
    'Weather', 'NPCs', 'Quests', 'FATEs', 'Levequests', 'Hunting Log Targets',
    'Monsters', 'Logging', 'Harvesting', 'Mining', 'Quarrying',
    'Fishing Log Locations', "Open this page in Garland Data",
    "Open this page in the Eorzea Database", "Run a search in YouTube",
    "Tank", "Healer", "Damage Dealer", "Trust System", "Information Needed",
    "Add Image",
}
TITLE_REPLACEMENTS = {
    "#Rewarded from TT Duel": "",
    "#Acquired from Chest": "",
    " (Item)": "",
    "#Available during Event": "",
    "#Received from Coffer": "",
    "Torn from the Heavens-The Dark Colossus Destroys All": "Torn from the Heavens/The Dark Colossus Destroys All",
}

def load_workbook_from_url(url):
    file = requests.get(url, timeout=60)
    file.raise_for_status()
    return load_workbook(filename=BytesIO(file.content))


def read_xlsx_file():
    KEY = os.environ.get("GDRIVE_APIKEY")
    MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    SHEET = "11SBQWYyFnyh19Ku6T1Wr5tNAJvxdze8tvD6m4bzPGIA"
    r = f"https://www.googleapis.com/drive/v3/files/{SHEET}/export?key={KEY}&mimeType={MIME_TYPE}"
    wb = load_workbook_from_url(r)
    #wb = openpyxl.load_workbook('./guide_ffxiv.xlsx')
    sheet = wb['Tabelle1']
    max_row = sheet.max_row
    max_column = sheet.max_column
    return sheet, max_row, max_column


def get_items_per_fight(name):
    url = f"https://ffxiv.gamerescape.com/wiki/{name}"
    page = requests.get(url, timeout=60)
    if page.status_code == 200:
        content = page.content
    else:
        print_color_red(url)
        return []
    print(url)
    DOMdocument = BeautifulSoup(content, 'html.parser')
    items = []
    seen_items = set()
    for table in DOMdocument.find_all('table'):
        #print(table)
        for a in table.find_all('a'):
            if a.find_all("img"):
                title = a.get("title")
                if title is None:
                    continue
                if any(skip in title for skip in SKIP_TITLE_PARTS):
                    continue
                if title in SKIP_TITLES:
                    continue
                for old, new in TITLE_REPLACEMENTS.items():
                    title = title.replace(old, new)

                item_id = enitem_lookup.get(title.lower())
                if item_id and item_id not in seen_items:
                    items.append(item_id)
                    seen_items.add(item_id)
    return items


def get_data_from_xlsx(sheet, max_column, i):
    return str(sheet.cell(row=int(i), column=int(29)).value).replace("None", "").strip()


def run():
    sheet, max_row, max_column = read_xlsx_file()
    #max_row = 20
    result = ""
    f_resutl = {}
    for i in range(2, max_row):
        entry = get_data_from_xlsx(sheet, max_column, i)
        if entry == "":
            continue
        items = get_items_per_fight(entry)
        entry_value = f'\t"{entry}": {items}' + ",\n"
        if entry_value not in result:
            result += entry_value
            f_resutl[entry] = items


    print(f_resutl)
    writeJsonFile("after_item_scan.json", f_resutl)


def test():
    befor = readJsonFile("before_item_scan.json")
    after = readJsonFile("after_item_scan.json")
    delete_after = []
    for key, value in after.items():
        for item in value:
            try:
                befor[key].remove(item)
            except Exception as e:
                print(e)
    print_pretty_json(befor)

if __name__ == "__main__":
    run()
