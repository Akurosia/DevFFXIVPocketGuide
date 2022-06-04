#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# coding: utf8
import copy
import os
from operator import itemgetter
import traceback
# traceback.print_exc()
import re
import errno
import openpyxl
import yaml
from yaml.loader import SafeLoader
import math
import natsort
from collections import OrderedDict
import convert_skills_to_guide_form as csgf
import generateLinks as gl
import logging

import sys
from ffxiv_aku import *

import logging

enemy = {
    "title": "",
    "title_en": "",
    "id": "",
    "hp": {
        "min": 100,
        "max": 0
    },
    "attacks": [{
        "title": "",
        "title_id": "",
        "title_en": "",
        "attack_in_use": "",
        "disable": "",
        "type": "",
        "damage_type": "",
        "damage": {
            "min": 100,
            "max": 0
        },
        "phases": [{"phase": "", }],
        "roles": [{"role": "", }],
        "tags": [{"tag": "", }],
        "notes": [{"note": "", }],
    }],
}
example_sequence = {
    "sequence": [{
        "phase": "09",
        "name": "phase_name",
        "alerts": [{
            "alert": "Die folgenden Angriffe haben sind entweder unbekannt oder haben keine klare Herkunft",
        }],
        "mechanics": [{
            "title": "sequence-mechanic-01",
            "notes": [{
                "note": "sequence-mechanic-note-01",
            }],
        }],
        "attacks": [{
            "attack": "sequence-attack-01",
        }],
        "images": [{
            "url": "/assets/img/test.jpg",
            "alt": "/assets/img/test.jpg",
            "height": "250px",
        }],
        "videos": [{
            "url": "https&#58;//ffxivguide.akurosia.de/upload/test.mp4",
        }],
    }]
}
example_add_sequence = {
    "sequence": [{"phase": "09", }]
}


class CustomFormatter(logging.Formatter):
    grey = "\x1b[38;20m"
    yellow = "\x1b[33;20m"
    red = "\x1b[31;20m"
    bold_red = "\x1b[31;1m"
    reset = "\x1b[0m"
    format = "%(asctime)s - %(name)s - %(levelname)s - %(message)s (%(filename)s:%(lineno)d)"
    FORMATS = {
        logging.DEBUG: wrap_in_color_green(format),
        logging.INFO: wrap_in_color_blue(format),
        logging.WARNING: wrap_in_color_yellow(format),
        logging.ERROR: wrap_in_color_red(format),
        logging.CRITICAL: bold_red + format + reset
    }

    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)


logger = logging.getLogger("My_app")
logger.setLevel(logging.CRITICAL)
ch = logging.StreamHandler()
ch.setLevel(logging.CRITICAL)
ch.setFormatter(CustomFormatter())
logger.addHandler(ch)

disable_green_print = True
disable_yellow_print = True
disable_blue_print = True
disable_red_print = True

storeFilesInTmp(True)
logdata = get_any_Logdata()
patchversions = get_any_Versiondata()
logdata_lower = dict((k.lower(), v) for k, v in logdata.items())
action = loadDataTheQuickestWay("action_all.json", translate=True)
territorytype = loadDataTheQuickestWay("territorytype_all.json", translate=True)
contentfindercondition = loadDataTheQuickestWay("contentfindercondition_all.json", translate=True)
contentfinderconditionX = loadDataTheQuickestWay("ContentFinderCondition.de.json")
contentmembertype = loadDataTheQuickestWay("ContentMemberType.json")
quests = loadDataTheQuickestWay("Quest.de.json")
levels = loadDataTheQuickestWay("Level.json")
maps = loadDataTheQuickestWay("Map.json")
placename = loadDataTheQuickestWay("placename_all.json", translate=True)
bnpcname = loadDataTheQuickestWay("bnpcname_all.json", translate=True)
eobjname = loadDataTheQuickestWay("eobjname_all.json", translate=True)
status = loadDataTheQuickestWay("status_all.json", translate=True)
enpcresident = loadDataTheQuickestWay("enpcresident_all.json", translate=True)
mounts = loadDataTheQuickestWay("mount_all.json", translate=True)
minions = loadDataTheQuickestWay("companion_all.json", translate=True)
orchestrions = loadDataTheQuickestWay("orchestrion_all.json", translate=True)
ttcards = loadDataTheQuickestWay("tripletriadcard_all.json", translate=True)
LANGUAGES = ["de", "en", "fr", "ja", "cn", "ko"]
XLSXELEMENTS = ["exclude", "date", "sortid", "title", "categories", "slug", "image", "patchNumber", "patchName", "difficulty", "plvl", "plvl_sync", "ilvl", "ilvl_sync", "quest_id", "gearset_loot", "tt_card1", "tt_card2", "orchestrion", "orchestrion2", "orchestrion3", "orchestrion4", "orchestrion5", "orchestrion_material1", "orchestrion_material2", "orchestrion_material3", "mtqvid1", "mtqvid2", "mrhvid1", "mrhvid2", "mount1", "mount2", "minion1", "minion2", "minion3", "instanceType", "mapid", "bosse", "adds", "mechanics", "tags", "teamcraftlink", "garlandtoolslink", "gamerescapelink", "done"]
UNKNOWNTITLE = {'de': 'Unbekannte Herkunft', 'en': 'Unknown Source', 'fr': 'Unknown Source', 'ja': 'Unknown Source', 'cn': 'Unknown Source', 'ko': 'Unknown Source'}


def read_xlsx_file():
    # open file, get sheet, last row and last coulmn
    wb = openpyxl.load_workbook('./guide_ffxiv.xlsx')
    sheet = wb['Tabelle1']
    max_row = sheet.max_row
    max_column = sheet.max_column
    return sheet, max_row, max_column


def getPrevAndNextContentOrder(sheet, elements, max_row):
    entry = {}
    for i in range(1, max_row + 1):
        instanceType = str(sheet.cell(row=int(i), column=int(elements.index('instanceType')) + 1).value).replace("None", "")
        if not entry.get(instanceType, None):
            entry[instanceType] = {}
        sortID = str(sheet.cell(row=int(i), column=int(3)).value).replace("None", "")
        addon = str(sheet.cell(row=int(i), column=int(5)).value).replace("None", "")
        slug = str(sheet.cell(row=int(i), column=int(6)).value).replace("None", "")
        entry[instanceType][sortID] = "/" + addon + "/" + slug
    return OrderedDict(natsort.natsorted(entry.items()))


def get_data_from_xlsx(sheet, max_column, i, elements):
    entry = {}
    # for every column in row add all elements into a dict:
    # max_column will ignore last column due to how range is working
    for j in range(1, max_column + 1):
        entry[elements[j - 1]] = str(sheet.cell(row=int(i), column=int(j)).value).replace("None", "")
    return entry


def clean_entries_from_single_quotes(entry):
    for key, value in entry.items():
        if value.startswith("'"):
            entry[key] = value[1:]
        if value.endswith("'"):
            entry[key] = value[:-1]
    return entry


def getMaps(_map):
    global maps
    for key, value in maps.items():
        if value['Id'] == _map:
            return value
    sys.exit()


def truncate(f, n):
    # line to collaps
    return math.floor(f * 10 ** n) / 10 ** n


def getLevel(level):
    global levels
    level = levels[level.replace('Level#', "")]
    map_ = getMaps(level['Map'])
    x = truncate(ToMapCoordinate(float(level['X'].replace(",", ".")), float(map_['SizeFactor'])), 1)
    y = truncate(ToMapCoordinate(float(level['Z'].replace(",", ".")), float(map_['SizeFactor'])), 1)
    return {"x": x, "y": y, "region": map_['PlaceName']['Region'], "placename": map_['PlaceName']['Value']}


def ToMapCoordinate(val, mapsize):
    c = mapsize / 100.0
    val *= c
    return ((41.0 / c) * ((val + 1024.0) / 2048.0)) + 1


def workOnQuests(entry, quest_id):
    global quests
    if quest_id == "":
        entry['quest'] = ""
        entry['quest_location'] = ""
        entry['quest_npc'] = ""
        return entry
    quest = quests[quest_id]
    entry['quest'] = quest['Name'].replace(" ", "").replace(" ", "")
    entry['quest_npc'] = quest['Issuer']['Start']
    try:
        level_data = getLevel(quest['Issuer']['Location'])
        entry['quest_location'] = f'{level_data["placename"]} ({level_data["x"]}, {level_data["y"]})'
    except KeyError:
        entry['quest_location'] = ""
        print_color_red(f"[workOnQuests] Error on loading: {quest['Issuer']['Location']} ({quest_id})")
    return entry


def seperate_data_into_array(tag, entry):
    if entry[tag]:
        entry[tag] = entry[tag].strip("'[").strip("]'").strip("\"[").strip("]\"").strip("[").strip("]").replace("\", \"", "', '").replace("\",\"", "', '").split("', '")
        entry[tag] = [b for b in entry[tag]]


def uglyContentNameFix(name, instanceType=None, difficulty=None):
    if difficulty == "Fatal" and instanceType == "ultimate" and "fatal" not in name.lower():
        name = f"{name} (fatal)"
    elif difficulty == "Episch" and instanceType == "raid" and "episch" not in name.lower():
        name = f"{name} (episch)"
    elif difficulty == "Episch" and instanceType == "feldexkursion" and "episch" not in name.lower():
        name = f"{name} (episch)"
    elif difficulty == "Schwer" and instanceType == "dungeon" and "schwer" not in name.lower():
        name = f"{name} (schwer)"
    # handle stupid edge cases for primals
    elif name in ["Königliche Konfrontation", "Jagd auf Rathalos"] and difficulty.lower() != "normal":
        name = f"{name} ({difficulty.lower()})"
    elif name in ["Memoria Misera"] and difficulty.lower() != "normal":
        name = f"{name} ({difficulty.lower()})"
    elif name in ["Krieger des Lichts"] and difficulty.lower() == "extrem":
        name = f"{name} ({difficulty.lower()})"
    # handle edge cases PvP
    elif "(Flechtenhain)" in name:
        name = name.replace("(Flechtenhain)", "")
    elif "(Kampfplatz)" in name:
        name = name.replace("(Kampfplatz)", "")
    # placeholder
    elif name == "":
        return ""
    # make sure brackets are always lowercase
    name = name.replace("(Fatal)", "(fatal)").replace("(Episch)", "(episch)").replace("(Schwer)", "(schwer)").replace("(Extrem)", "(extrem)")
    return name


def getContentName(name, lang="en", difficulty=None, instanceType=None):
    name = uglyContentNameFix(name, instanceType, difficulty)
    try:
        for key, content in contentfindercondition.items():
            if "memoria" in content["Name_de"].lower().strip() and "memoria" in name.lower().strip():
                return content[f"Name_{lang}"]
            if content["Name_de"].lower().strip() == name.lower().strip():
                return content[f"Name_{lang}"]
        for key, place in placename.items():
            if place["Name_de"].lower().strip() == name.lower().strip():
                return place[f"Name_{lang}"]
    except KeyError:
        pass
    if name not in ['title']:
        print_color_red("Could not translate: " + name)
    return ""


def getEntriesForRouletts(entry):
    global contentfinderconditionX
    for key, value in contentfinderconditionX.items():
        if value['Name'] == getContentName(entry["title"], "de", entry["difficulty"], entry["instanceType"]):
            entry['type'] = value['ContentType'].lower()
            entry['mapid'] = value['TerritoryType']
            entry['allianceraid'] = value['AllianceRoulette']
            entry['frontier'] = value['FeastTeamRoulette']
            entry['expert'] = value['ExpertRoulette']
            entry['guildhest'] = value['GuildHestRoulette']
            entry['level50_60_70'] = value['Level50/60/70Roulette']
            entry['level80'] = value['Level80Roulette']
            entry['leveling'] = value['LevelingRoulette']
            entry['main'] = value['MSQRoulette']
            entry['mentor'] = value['MentorRoulette']
            entry['normalraid'] = value['NormalRaidRoulette']
            entry['trial'] = value['TrialRoulette']
            return entry
    return entry


def getEntryData(sheet, max_column, i, elements, orderedContent):
    entry = get_data_from_xlsx(sheet, max_column, i, elements)
    entry = clean_entries_from_single_quotes(entry)
    entry = workOnQuests(entry, entry["quest_id"])
    entry = getEntriesForRouletts(entry)
    for lang in LANGUAGES:
        entry[f"title_{lang}"] = getContentName(entry["title"], lang, entry["difficulty"], entry["instanceType"])
    _previous, _next = getBeforeAndAfterContentEntries(orderedContent, entry)
    # remove time from excel datetime
    entry["date"] = str(entry["date"]).replace(" 00:00:00", "").replace("-", ".")
    entry["prev_content"] = _previous
    entry["next_content"] = _next
    entry["line_index"] = i
    seperate_data_into_array("bosse", entry)
    seperate_data_into_array("adds", entry)
    seperate_data_into_array("tags", entry)
    return entry


def get_old_content_if_file_is_found(_existing_filename):
    if os.path.exists(_existing_filename):
        with open(_existing_filename, encoding="utf8") as f:
            doc = list(yaml.load_all(f, Loader=SafeLoader))[0]
            return doc
    return {}


def getBeforeAndAfterContentEntries(orderedContent, entry):
    _previous = None
    _next = None
    _type = orderedContent[entry['instanceType']]
    _typeKeys = list(_type)
    for i, k in enumerate(_type):
        if _type[k].endswith(entry['slug']):
            if i - 1 >= 0:
                try:
                    _previous = _type[_typeKeys[i - 1]]
                except:
                    pass
            try:
                _next = _type[_typeKeys[i + 1]]
            except:
                pass
            return _previous, _next
    return None, None


def try_to_create_file(filename):
    if not os.path.exists(os.path.dirname(filename)):
        try:
            os.makedirs(os.path.dirname(filename))
        except OSError as exc:  # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise


def get_territorytype_from_mapid(mapid):
    for key, tt_type in territorytype.items():
        if tt_type["TerritoryType"].lower() == mapid.lower():
            return tt_type
    print_color_red(f"Could not find territorytype for {mapid}")
    return ""


def replaceSlug(text):
    return str(text).replace("_", "-").replace(".", "-").replace(",", "").replace("'", "").replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("Ä", "Ae").replace("Ö", "Oe").replace("Ü", "Ue").replace("ß", "ss")


def getImage(image):
    image = image.replace(".tex", "_hr1.png\"")
    image = image.replace("ui/icon/", "")
    return image


def checkVariable(element, name):
    if element[name] and not element[name] == "###":
        return True
    return False


def get_order_id(entry):
    patch = "0000"
    plvl = "00"
    sortid = "0000"
    if "." in entry["patchNumber"]:
        t_minor = 0
        major, minor = entry["patchNumber"].split(".")
        if str(minor).endswith("a") or str(minor).endswith("b"):
            t_minor = int(minor[:-1])
            t_minor = t_minor * 10 if t_minor < 10 else t_minor
            t_minor = t_minor * 10 if t_minor < 100 else t_minor
            if str(minor).endswith("a"):
                minor = t_minor + 1
            elif str(minor).endswith("b"):
                minor = t_minor + 2
        else:
            minor = f"{minor}0" if len(minor) < 2 else minor
            minor = f"{minor}0" if len(minor) < 3 else minor
        patch = f"{major}{minor}"
    if entry["plvl"]:
        plvl = int(entry["plvl"]) * 10 if len(entry["plvl"]) < 2 else entry["plvl"]
    if entry["sortid"]:
        sortid = f"0{entry['sortid']}" if len(entry["sortid"]) < 1 else entry["sortid"]
        sortid = f"0{sortid}" if len(sortid) < 2 else sortid
        sortid = f"0{sortid}" if len(sortid) < 3 else sortid
        sortid = f"0{sortid}" if len(sortid) < 4 else sortid
    return f"{patch}{plvl}{sortid}"


def addEntries(header_data, entry, field, get_data_function):
    if checkVariable(entry, field):
        header_data += '  - name: "' + entry[field] + '"\n'
        mount_id = get_data_function(entry[field])
        if mount_id:
            header_data += '    id: "' + mount_id + '"\n'
    return header_data


def getMountIDByName(name):
    for _id, mount in mounts.items():
        if mount['Singular_de'] == name:
            return _id
    return None


def getMinionIDByName(name):
    for _id, minion in minions.items():
        if minion['Singular_de'] == name:
            return _id
    return None


def getOrchestrionIDByName(name):
    for _id, orchestrion in orchestrions.items():
        if orchestrion['Name_de'].lower() == name.lower():
            return _id
    return None


def getTTCardIDByName(name):
    for id, ttcard in ttcards.items():
        if ttcard['Name_de'] == name:
            return id.split(".0")[0]
    return None


def get_video_url(url):
    if url.startswith("https"):
        return url
    return "https://www.youtube.com/watch?v={}".format(url)


def rewrite_content_even_if_exists(entry, old_wip):
    header_data = ""
    tt_type_name = get_territorytype_from_mapid(entry["mapid"])
    if old_wip in ["True", "False"]:
        header_data += 'wip: "' + str(old_wip).title() + '"\n'
    else:
        header_data += 'wip: "True"\n'
    #header_data += 'title: "' + entry["title"] + '"\n'
    header_data += 'title:\n'
    for lang in LANGUAGES:
        tmp = entry[f"title_{lang}"].replace(f' ({entry["difficulty"].lower()})', "")
        tmp = tmp.replace(f' ({entry["difficulty"].title()})', "")
        tmp = tmp.replace(f'Traumprüfung - ', "")
        header_data += f'  {lang}: "' + tmp + '"\n'
    header_data += 'layout: guide_post\n'
    header_data += 'page_type: guide\n'
    header_data += f'excel_line: \"{entry["line_index"]}\"\n'
    header_data += 'categories: "' + entry["categories"] + '"\n'
    header_data += 'patchNumber: "' + entry["patchNumber"].replace("'", "") + '"\n'
    if patchversions.get(entry["patchNumber"], None):
        header_data += 'patchLink: "' + patchversions[entry["patchNumber"]]['link_to_patch'] + '"\n'
    header_data += 'difficulty: "' + entry["difficulty"] + '"\n'
    header_data += 'instanceType: "' + entry["instanceType"] + '"\n'
    header_data += 'date: "' + entry["date"] + '"\n'
    header_data += 'slug: "' + replaceSlug(entry["slug"]) + '"\n'
    if entry["prev_content"]:
        header_data += 'previous_slug: "' + replaceSlug(entry["prev_content"]) + '"\n'
    if entry["next_content"]:
        header_data += 'next_slug: "' + replaceSlug(entry["next_content"]) + '"\n'
    if entry["image"]:
        header_data += 'image:\n'
        header_data += '  - url: \"/' + getImage(entry["image"]) + '\n'
        #header_data += '    url: \"/' + getImage(entry["image"]) + '\n'
    header_data += 'terms:\n'
    header_data = writeTags(header_data, entry, tt_type_name)
    header_data += 'patchName: "' + entry["patchName"] + '"\n'
    if entry.get("mapid", None):
        header_data += 'mapid: "' + entry["mapid"] + '"\n'
    if not tt_type_name == "":
        header_data += 'contentname: "' + tt_type_name["Name_de"] + '"\n'
    header_data += 'sortid: ' + entry["sortid"] + '\n'
    header_data += 'plvl: ' + entry["plvl"] + '\n'
    header_data += 'plvl_sync: ' + entry["plvl_sync"] + '\n'
    header_data += 'ilvl: ' + entry["ilvl"] + '\n'
    header_data += 'ilvl_sync: ' + entry["ilvl_sync"] + '\n'
    if not entry["quest"] == "":
        header_data += 'quest: "' + entry["quest"] + '"\n'
    if not entry["quest_location"] == "":
        header_data += 'quest_location: "' + entry["quest_location"] + '"\n'
    if not entry["quest_npc"] == "":
        header_data += 'quest_npc: "' + entry["quest_npc"] + '"\n'
    header_data += 'order: ' + get_order_id(entry) + '\n'
    # mounts
    if checkVariable(entry, "mount1") or checkVariable(entry, "mount2"):
        header_data += 'mount:\n'
        header_data = addEntries(header_data, entry, "mount1", getMountIDByName)
        header_data = addEntries(header_data, entry, "mount2", getMountIDByName)
    # minions
    if checkVariable(entry, "minion1") or checkVariable(entry, "minion2") or checkVariable(entry, "minion3"):
        header_data += 'minion:\n'
        header_data = addEntries(header_data, entry, "minion1", getMinionIDByName)
        header_data = addEntries(header_data, entry, "minion2", getMinionIDByName)
        header_data = addEntries(header_data, entry, "minion3", getMinionIDByName)
    # gearset_loot
    if checkVariable(entry, "gearset_loot"):
        header_data += 'gearset_loot:\n'
        for gset in entry["gearset_loot"] .split(","):
            header_data += '  - gsetname: "' + gset + '"\n'
    # tt_cards
    if checkVariable(entry, "tt_card1") or checkVariable(entry, "tt_card2"):
        header_data += 'tt_card:\n'
        header_data = addEntries(header_data, entry, "tt_card1", getTTCardIDByName)
        header_data = addEntries(header_data, entry, "tt_card2", getTTCardIDByName)
    # orchestrion
    if checkVariable(entry, "orchestrion") or checkVariable(entry, "orchestrion2") or checkVariable(entry, "orchestrion3") or checkVariable(entry, "orchestrion4") or checkVariable(entry, "orchestrion5"):
        header_data += 'orchestrion:\n'
        header_data = addEntries(header_data, entry, "orchestrion", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion2", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion3", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion4", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion5", getOrchestrionIDByName)
    # orchestrion material
    if checkVariable(entry, "orchestrion_material1") or checkVariable(entry, "orchestrion_material2") or checkVariable(entry, "orchestrion_material3"):
        header_data += 'orchestrion_material:\n'
        if checkVariable(entry, "orchestrion_material1"):
            header_data += '  - name: "' + entry["orchestrion_material1"] + '"\n'
        if checkVariable(entry, "orchestrion_material2"):
            header_data += '  - name: "' + entry["orchestrion_material2"] + '"\n'
        if checkVariable(entry, "orchestrion_material3"):
            header_data += '  - name: "' + entry["orchestrion_material3"] + '"\n'
    # rouletts
    if entry.get("expert", None):
        header_data += 'rouletts:\n'
        if entry["allianceraid"]:
            header_data += '    allianceraid: ' + entry["allianceraid"] + "\n"
        if entry["frontier"]:
            header_data += '    frontier: ' + entry["frontier"] + "\n"
        if entry["expert"]:
            header_data += '    expert: ' + entry["expert"] + "\n"
        if entry["guildhest"]:
            header_data += '    guildhest: ' + entry["guildhest"] + "\n"
        if entry["level50_60_70"]:
            header_data += '    level50_60_70: ' + entry["level50_60_70"] + "\n"
        if entry["level80"]:
            header_data += '    level80: ' + entry["level80"] + "\n"
        if entry["leveling"]:
            header_data += '    leveling: ' + entry["leveling"] + "\n"
        if entry["main"]:
            header_data += '    main: ' + entry["main"] + "\n"
        if entry["mentor"]:
            header_data += '    mentor: ' + entry["mentor"] + "\n"
        if entry["normalraid"]:
            header_data += '    normalraid: ' + entry["normalraid"] + "\n"
        if entry["trial"]:
            header_data += '    trial: ' + entry["trial"] + "\n"
    # links:
    if checkVariable(entry, "teamcraftlink") or checkVariable(entry, "garlandtoolslink") or checkVariable(entry, "gamerescapelink"):
        header_data += 'links:\n'
        if checkVariable(entry, "teamcraftlink"):
            header_data += f'    teamcraftlink: "' + entry["teamcraftlink"] + '"\n'
        if checkVariable(entry, "garlandtoolslink"):
            header_data += f'    garlandtoolslink: "' + entry["garlandtoolslink"] + '"\n'
        if checkVariable(entry, "gamerescapelink"):
            header_data += f'    gamerescapelink: "' + entry["gamerescapelink"] + '"\n'
    # videos
    if checkVariable(entry, "mtqvid1"):
        header_data += 'mtq_vid1: "' + get_video_url(entry["mtqvid1"]) + '"\n'
    if checkVariable(entry, "mtqvid2"):
        header_data += 'mtq_vid2: "' + get_video_url(entry["mtqvid2"]) + '"\n'
    if checkVariable(entry, "mrhvid1"):
        header_data += 'mrh_vid1: "' + get_video_url(entry["mrhvid1"]) + '"\n'
    if checkVariable(entry, "mrhvid2"):
        header_data += 'mrh_vid2: "' + get_video_url(entry["mrhvid2"]) + '"\n'
    return header_data, entry


def writeTags(header_data, entry, tt_type_name):
    # write tags per expansion
    if entry["categories"] == "arr":
        header_data += "  - term: \"A Realm Reborn\"\n"
        header_data += "  - term: \"ARR\"\n"
    elif entry["categories"] == "hw":
        header_data += "  - term: \"Heavensward\"\n"
        header_data += "  - term: \"HW\"\n"
    elif entry["categories"] == "sb":
        header_data += "  - term: \"Stormblood\"\n"
        header_data += "  - term: \"SB\"\n"
    elif entry["categories"] == "shb":
        header_data += "  - term: \"Shadowbringers\"\n"
        header_data += "  - term: \"ShB\"\n"
    elif entry["categories"] == "ew":
        header_data += "  - term: \"Endwalker\"\n"
        header_data += "  - term: \"EW\"\n"
    else:
        pass

    if not tt_type_name == "":
        for lang in LANGUAGES:
            header_data += "  - term: \"" + tt_type_name["Name_" + lang] + "\"\n"

    for lang in LANGUAGES:
        header_data += "  - term: \"" + entry[f"title_{lang}"] + "\"\n"

    # write rest of the tags
    header_data += "  - term: \"" + entry["difficulty"] + "\"\n"
    header_data += "  - term: \"" + entry["patchNumber"] + "!\"\n"
    header_data += "  - term: \"" + entry["patchName"] + "\"\n"
    if not entry.get("quest", "") == "":
        header_data += "  - term: \"" + entry["quest"] + "\"\n"
    if checkVariable(entry, "mount1") or checkVariable(entry, "mount2"):
        header_data += "  - term: \"mounts\"\n"
        header_data += "  - term: \"Reittier\"\n"
    if checkVariable(entry, "minion1") or checkVariable(entry, "minion2") or checkVariable(entry, "minion3"):
        header_data += "  - term: \"minions\"\n"
        header_data += "  - term: \"Begleiter\"\n"
    if checkVariable(entry, "tt_card1") or checkVariable(entry, "tt_card2"):
        header_data += "  - term: \"tt_cards\"\n"
        header_data += "  - term: \"Triple Triad Karte\"\n"
    if checkVariable(entry, "gearset_loot"):
        for gset in entry["gearset_loot"].split(","):
            header_data += "  - term: \"" + gset + "\"\n"
    if checkVariable(entry, "orchestrion") or checkVariable(entry, "orchestrion2") or checkVariable(entry, "orchestrion3") or checkVariable(entry, "orchestrion4") or checkVariable(entry, "orchestrion5"):
        header_data += "  - term: \"orchestrion\"\n"
        header_data += "  - term: \"Notenrolle\"\n"
    if checkVariable(entry, "orchestrion_material1") or checkVariable(entry, "orchestrion_material2") or checkVariable(entry, "orchestrion_material3"):
        header_data += "  - term: \"orchestrion_material\"\n"
    if entry["instanceType"] == "trial":
        header_data += "  - term: \"Prüfung\"\n"
        header_data += "  - term: \"Trial\"\n"
        header_data += "  - term: \"Primae\"\n"
        header_data += "  - term: \"Primal\"\n"
    header_data += "  - term: \"" + entry["instanceType"] + "\"\n"

    found_roulette = False
    if entry.get("allianceraid", None) == "True":
        header_data += "  - term: \"allianceraid\"\n"
        found_roulette = True
    if entry.get("frontier", None) == "True":
        header_data += "  - term: \"frontier\"\n"
        found_roulette = True
    if entry.get("expert", None) == "True":
        header_data += "  - term: \"expert\"\n"
        found_roulette = True
    if entry.get("guildhest", None) == "True":
        header_data += "  - term: \"guildhest\"\n"
        found_roulette = True
    if entry.get("level50_60_70", None) == "True":
        header_data += "  - term: \"level50_60_70\"\n"
        found_roulette = True
    if entry.get("level80", None) == "True":
        header_data += "  - term: \"level80\"\n"
        found_roulette = True
    if entry.get("leveling", None) == "True":
        header_data += "  - term: \"leveling\"\n"
        found_roulette = True
    if entry.get("main", None) == "True":
        header_data += "  - term: \"main\"\n"
        found_roulette = True
    if entry.get("mentor", None) == "True":
        header_data += "  - term: \"mentor\"\n"
        found_roulette = True
    if entry.get("normalraid", None) == "True":
        header_data += "  - term: \"normalraid\"\n"
        found_roulette = True
    if entry.get("trial", None) == "True":
        header_data += "  - term: \"trial\"\n"
        found_roulette = True
    if found_roulette:
        header_data += "  - term: \"Zufallsinhalt\"\n"
        header_data += "  - term: \"roulette\"\n"

    if not entry["bosse"] == ['']:
        for b in entry["bosse"]:
            if b != "Unknown_":
                header_data += "  - term: \"" + b + "\"\n"
    if not entry["tags"] == ['']:
        for t in entry["tags"]:
            if t != "Unknown_":
                header_data += "  - term: \"" + t + "\"\n"
    return header_data


def cleanup_logdata(logdata_instance_content):
    try:
        del logdata_instance_content["combatants"]
    except Exception:
        pass
    try:
        del logdata_instance_content["zone"]
    except Exception:
        pass
    try:
        del logdata_instance_content["contentzoneid"]
    except Exception:
        pass
    music = None
    try:
        music = logdata_instance_content["music"]
        del logdata_instance_content["music"]
    except Exception:
        pass
    for enemy_name, enemy in logdata_instance_content.items():
        # try: del enemy["status"]
        # except Exception: pass
        try:
            del enemy["tether"]
        except Exception:
            pass
        try:
            del enemy["headmarker"]
        except Exception:
            pass
    new_lic = {}
    for k, v in logdata_instance_content.items():
        # if not k == "":
        new_lic[k] = v
    return new_lic, music


def getDataFromLogfile(entry):
    logdata_instance_content = None
    music = None
    contentzoneid = ""
    # get correct title capitalization to read data from logdata
    title = uglyContentNameFix(entry["title_de"].title(), entry["instanceType"], entry["difficulty"])
    # get the latest data from logdata
    if not entry["title_de"] == "" and logdata_lower.get(entry["title_de"].lower()):
        try:
            logdata_instance_content = dict(logdata[getContentName(title, lang="de")])
        except Exception:
            logdata_instance_content = dict(logdata[title])
        if logdata_instance_content.get('contentzoneid', None):
            contentzoneid = logdata_instance_content['contentzoneid']
        logdata_instance_content, music = cleanup_logdata(logdata_instance_content)
    return logdata_instance_content, music, contentzoneid


def addContentZoneIdToHeader(header_data, contentzoneid, entry):
    global contentfinderconditionX
    cmt = None
    if not contentzoneid == "":
        header_data += 'contentzoneids:\n'
        for zone in contentzoneid:
            header_data += '  - id: ' + zone + '\n'
    for key, value in contentfinderconditionX.items():
        if value['Name'] == entry['title_de']:
            cmt = value['ContentMemberType']
            if not "InstanceContent" in value['Content']:
                continue
            contentid = value['Content'].replace("InstanceContent#", "")
            if not contentid:
                continue
            _id = "8003" + str(hex(int(contentid))[2:]).rjust(4, '0').upper()
            if "contentzoneids:" not in header_data:
                header_data += 'contentzoneids:\n'

            if _id not in header_data:
                # if _id not in contentzoneid:
                header_data += '  - id: ' + _id + '\n'
    return header_data, cmt


def addGroupCollections(cmt, entry):
    global contentmembertype
    header_data = ""
    skip_lookoup = False
    if not cmt:
        if "Traumprüfung" in entry['title'] or "Dalriada" in entry['title'] or "Castrum Lacus Litore" in entry['title']:
            cmt_entry = 8
            healerp = 2
            tankp = 2
            meleep = 2
            rangep = 2
            skip_lookoup = true
        else:
            print("Could not find GroupCollection for: " + entry['title'])
            return header_data

    if not skip_lookoup:
        wanted_id = cmt.split("#")[1]
        cmt_entry = contentmembertype[f"{wanted_id}"]
        healerp = cmt_entry['HealersPerParty']
        tankp = cmt_entry['TanksPerParty']
        meleep = cmt_entry['MeleesPerParty']
        rangep = cmt_entry['RangedPerParty']

    if int(healerp) + int(tankp) + int(meleep) + int(rangep) > 0:
        header_data += "group:\n"
        if not healerp == "0":
            header_data += f'    healer: "{healerp}"\n'
        if not tankp == "0":
            header_data += f'    tank: "{tankp}"\n'
        if not meleep == "0":
            header_data += f'    melee: "{meleep}"\n'
        if not rangep == "0":
            header_data += f'    range: "{rangep}"\n'
    return header_data


def addMusic(header_data, music):
    header_data += "music:\n"
    for m in music:
        header_data += f"    - name: \"{m}\"\n"
        _id = getOrchestrionIDByName(m)
        if _id:
            header_data += f"      id: \"{_id}\"\n"
    return header_data


def delete_invalid_entries(tmp_attack):
    try:
        del tmp_attack['title_id']
    except Exception:
        pass
    try:
        del tmp_attack['damage_type']
    except Exception:
        pass
    try:
        del tmp_attack['roles']
    except Exception:
        pass
    try:
        del tmp_attack['tags']
    except Exception:
        pass
    return tmp_attack


def writeFileIfNoDifferent(filename, filedata):
    try:
        with open(filename, "r", encoding="utf8") as f:
            x_data = f.read()
    except:
        x_data = None

    if not filedata == x_data:
        with open(filename, "w", encoding="utf8") as fi:
            fi.write(filedata)
        print(f"Wrote new data to file {filename}")


# EVERYTHING RELATED TO GUIDE DATA
####################################################################################################################

def check_Mechanics(entry, old_mechanics):
    guide_data = ""
    # this contruct looks ugly but it forbidds the recreation of empty mechanics
    if old_mechanics:
        guide_data += "mechanics:\n"
        if old_mechanics:
            for mechanic in old_mechanics:
                guide_data += add_Mechanic(mechanic)
    # if no old mechanics are found and excel contains [] it will create the example mechanic
    elif entry["mechanics"] != "":
        guide_data += "mechanics:\n"
        mechanics = entry["mechanics"].strip("\"[").strip("]\"").split("\",\"")
        counter = 1
        for m in mechanics:
            mechanic = {
                "steps": [{
                    "step": "09",
                    "notes": [{"note": f"Mechanics-note {counter}", }],
                    "images": [{"url": "/assets/img/test.jpg", "alt": "/assets/img/test.jpg", "height": "250px", }],
                    "videos": [{"url": "https&#58;//akurosia.de/upload/test.mp4", }],
                }],
            }
            if mechanics == ['']:
                mechanic['title'] = f"Mechanic-Title {counter}"
            else:
                mechanic['title'] = m

            guide_data += add_Mechanic(mechanic)
            counter += 1
    return guide_data


def add_Mechanic(data):
    guide_data = ""
    guide_data += f"  - title: \"{data['title']}\"\n"
    if data.get("steps", None):
        guide_data += "    steps:\n"
        for step in data["steps"]:
            guide_data += f"      - step: \"{int(step['step']):02d}\"\n"

            if step.get("notes", None):
                guide_data += "        notes:\n"
                for note in step["notes"]:
                    guide_data += f"          - note: \"{note['note']}\"\n"

            if step.get("images", None):
                guide_data += "        images:\n"
                for image in step["images"]:
                    guide_data += f"          - url: \"{image['url']}\"\n"
                    guide_data += f"            alt: \"{image.get('alt', image['url'])}\"\n"
                    guide_data += f"            height: \"{image.get('height', '250px')}\"\n"

            if step.get("videos", None):
                guide_data += "        videos:\n"
                for video in step["videos"]:
                    guide_data += f"          - url: \"{video['url']}\"\n"
    return guide_data


def compare_skill_ids(old_enemy_data, new_enemy_data, existing_attacks, remove_attack):
    for attack_id in new_enemy_data.get('skill', {}):
        for attack in old_enemy_data.get('attacks', []):
            if type(attack['title']) == str:
                _id = attack.get('title_id', None)
                if not attack.get('title_id', None):
                    if attack.get('variation', None):
                        _id = attack['variation'][0]['title_id']
                attack['title'] = addLanguageElements("action", _id, attack['title'])
            existing_attacks[attack['title']['de']] = attack['type']
            if attack_id == attack.get('title_id', None):
                remove_attack.append(attack_id)
            if attack.get("variation", None):
                for vari in attack.get("variation", None):
                    if attack_id == vari['title_id']:
                        remove_attack.append(attack_id)
    return existing_attacks, remove_attack


def remove_skills_from_list_if_found(remove_attack, new_enemy_data):
    for x in remove_attack:
        try:
            del new_enemy_data["skill"][x]
        except Exception:
            pass
    return new_enemy_data


def addSortName(old_enemy_data, type_):
    if old_enemy_data.get(type_, None):
        for i, attack in enumerate(old_enemy_data[type_]):
            if not old_enemy_data[type_][i].get('sortname', None):
                if type(old_enemy_data[type_][i]['title']) == str:
                    x = "status"
                    if type_ == "attacks":
                        x = "action"
                    if not old_enemy_data[type_][i].get('title_id', None):
                        if old_enemy_data[type_][i].get('variation', None):
                            _id = old_enemy_data[type_][i]['variation'][0]['title_id']
                    else:
                        _id = old_enemy_data[type_][i]['title_id']
                    old_enemy_data[type_][i]['title'] = addLanguageElements(x, _id, old_enemy_data[type_][i]['title'])
                old_enemy_data[type_][i]['sortname'] = old_enemy_data[type_][i]['title']['de']
    return old_enemy_data


def sort_list_of_skills(data):
    if data.get('attacks', None):
        data['attacks'] = sorted(data['attacks'], key=itemgetter('sortname'))
        unknown_attacks = []
        known_attacks = []
        for attack in data['attacks']:
            if attack['title']['de'].startswith("Unknown_"):
                unknown_attacks.append(attack)
            else:
                known_attacks.append(attack)
        data['attacks'] = unknown_attacks + known_attacks
    return data


def translateAttack(skill_id, _type=action, lang="en"):
    try:
        text = _type[str(int(skill_id, 16))][f"Name_{lang}"]
        text = fixCaptilaziationAndRomanNumerals(text)
        if text == "":
            text = f"Unknown_{skill_id}"
        return text
    except KeyError:
        return f"Unknown_{skill_id}"


def fixCaptilaziationAndRomanNumerals(text):
    text = text.title()
    text = re.sub(r" (Ii|Iii|IIi|Vi|Vii|Viii|Iv|Ix|Xi|Xii|Xiii|Xliii|Iii-E|Xxiv|013Bl|Xii\.)\. ", lambda x:  x.group(0).upper(), text)
    text = re.sub(r" (Ii|Iii|IIi|Vi|Vii|Viii|Iv|Ix|Xi|Xii|Xiii|Xliii|Iii-E|Xxiv|013Bl|Xii\.)$", lambda x:  x.group(0).upper(), text)
    text = re.sub(r" (Ii) ", lambda x:  x.group(0).upper(), text)
    text = re.sub(r"('[a-zA-Z])(?! )", lambda x:  x.group(0).upper(), text)
    text = re.sub(r"('[a-zA-Z]) ", lambda x:  x.group(0).lower(), text)
    text = text.replace('"', r'\"')
    text = text.replace("&#246;", "ö").replace("&#252;", "ü").replace("&#228;", "ä").replace("&#223;", "ß")
    return text


def get_fixed_status_description(_id):
    try:
        description = status[str(int(_id, 16))]['Description_de']
        description = re.sub('<UIForeground>[0-9A-F]{1,6}</UIForeground>', '', description)
        description = re.sub('<UIGlow>[0-9A-F]{1,6}</UIGlow>', '', description)
        description = description.replace("\n\n", "<br>")
        description = description.replace("\n", "<br>")
        return description
    except KeyError:
        return f"Unknown_{_id}"


def getBnpcNameFromID(_id, aname, nname, lang="en"):
    bnew_name = ""
    enew_name = ""
    ennew_name = ""
    if type(_id) == list:
        _id = _id[0]
    _id = str(_id)
    try:
        bnew_name = bnpcname[_id]["Singular_de"]
        m = re.search(nname, bnew_name, re.IGNORECASE)
        n = re.search(aname, bnew_name, re.IGNORECASE)
        if m or n:
            return bnpcname[_id][f"Singular_{lang}"]
    except Exception:
        pass
    try:
        enew_name = eobjname[_id]["Singular_de"]
        m = re.search(nname, enew_name, re.IGNORECASE)
        n = re.search(aname, enew_name, re.IGNORECASE)
        if m or n:
            return eobjname[_id][f"Singular_{lang}"]
    except Exception:
        pass
    try:
        ennew_name = enpcresident[_id]["Singular_de"]
        m = re.search(nname, ennew_name, re.IGNORECASE)
        n = re.search(aname, ennew_name, re.IGNORECASE)
        if m or n:
            return enpcresident[_id][f"Singular_{lang}"]
    except Exception:
        pass

    if "α" not in bnew_name and "β" not in bnew_name and "（仮）鎖" not in bnew_name:
        print_color_red(f"'{bnew_name}', '{enew_name}', '{ennew_name}' not found {aname} ({nname}) - ({_id})")
    return ""


def getBnpcName(name, _id, lang="en"):
    name = name.lower()
    aname = ""
    # take care of all the article cases
    name_match = re.search("^(der|die|das) ", name)
    if name_match:
        aname = name.replace("der ", r"(\[t\]|(der|die|das)) ").replace("die ", r"(\[t\]|(der|die|das)) ").replace("das ", r"(\[t\]|(der|die|das)) ")

    name_match = re.search(" (der|die|das)", name)
    if name_match:
        aname = name.replace(" der ", r" (\[t\]|(der|die|das)) ").replace(" die ", r" (\[t\]|(der|die|das)) ").replace(" das ", r" (\[t\]|(der|die|das)) ")

    if aname == "":
        aname = name

    # take care of the german gender cases
    nname = aname.replace("er ", r"(\[a\]|(e|es|er|en)) ").replace("es ", r"(\[a\]|(e|es|er|en)) ").replace("en ", r"(\[a\]|(e|es|er|en)) ").replace("e ", r"(\[a\]|(e|es|er|en)) ")
    if nname.endswith("e"):
        nname = nname[:-1] + r"(\[a\]|(e|es|er|en))"
    elif nname.endswith("en") or nname.endswith("es") or nname.endswith("er"):
        nname = nname[:-2] + r"(\[a\]|(e|es|er|en))"

    if not _id == "":
        resultname = getBnpcNameFromID(_id, aname, nname, lang)
        if not resultname == "":
            return fixCaptilaziationAndRomanNumerals(resultname)
    # check results agains bnpcname
    for key, bnpc in bnpcname.items():
        m = re.search(nname, bnpc["Singular_de"].lower())
        if m:
            return fixCaptilaziationAndRomanNumerals(bnpc[f"Singular_{lang}"])

    # check results agains eobjname
    for key, eobj in eobjname.items():
        m = re.search(nname, eobj["Singular_de"].lower())
        if m:
            return fixCaptilaziationAndRomanNumerals(eobj[f"Singular_{lang}"])

    # check results agains eobjname
    for key, eobj in enpcresident.items():
        m = re.search(nname, eobj["Singular_de"].lower())
        if m:
            return fixCaptilaziationAndRomanNumerals(eobj[f"Singular_{lang}"])
    # return anything, just in case
    print_color_yellow(f"Missing {name} examples where ({aname}) and ({nname})", disable_yellow_print)
    return ""


def merge_attacks(old_enemy_data, new_enemy_data, enemy_type):
    remove_attack = []
    existing_attacks = {}
    # comapre all skill ids
    existing_attacks, remove_attack = compare_skill_ids(old_enemy_data, new_enemy_data, existing_attacks, remove_attack)
    # remove skill ids if they were found before
    new_enemy_data = remove_skills_from_list_if_found(remove_attack, new_enemy_data)

    old_enemy_data = addSortName(old_enemy_data, "attacks")
    old_enemy_data = sort_list_of_skills(old_enemy_data)
    if not new_enemy_data.get('skill', None):
        return old_enemy_data
    # merge new keys
    for attack_id, attack in new_enemy_data['skill'].items():
        if attack["name"] == "":
            attack["name"] = "Unknown_" + attack_id
        if existing_attacks.get(attack['name'], None):
            # convert regular attk to variation
            if existing_attacks[attack['name']] == "regular":
                print_color_blue(f"\t\tNeed to convert: {attack['name']} ({attack_id})", disable_blue_print)
                # find attack
                attack_index = None
                for i, old_attack in enumerate(old_enemy_data['attacks']):
                    if old_attack.get('title', {}).get('de', None) == fixCaptilaziationAndRomanNumerals(attack['name']):
                        attack_index = i
                        break
                tmp_attack = old_enemy_data['attacks'][attack_index]
                tmp_attack['type'] = "variation"
                tmp_attack['notes'] = [{'note': 'Variation-note BIG'}]
                tmp = {
                    'title': {
                        'en': translateAttack(tmp_attack['title_id'], lang="en"),
                        'de': translateAttack(tmp_attack['title_id'], lang="de"),
                        'fr': translateAttack(tmp_attack['title_id'], lang="fr"),
                        'ja': translateAttack(tmp_attack['title_id'], lang="ja"),
                        'cn': translateAttack(tmp_attack['title_id'], lang="cn"),
                        'ko': translateAttack(tmp_attack['title_id'], lang="ko")
                    },
                    'title_id': tmp_attack['title_id'],
                    'damage_type': tmp_attack['damage_type'],
                    'single_or_aoe': attack['type_id']
                }
                if tmp_attack['title'] == "Attacke":
                    tmp['disable'] = 'false'
                if tmp_attack.get('damage', None):
                    tmp["damage"] = {}
                    try:
                        tmp["damage"]["min"] = tmp_attack['damage']['min']
                        tmp["damage"]["max"] = tmp_attack['damage']['max']
                    except:
                        tmp["damage"]["min"] = tmp_attack['damage'][0]['min']
                        tmp["damage"]["max"] = tmp_attack['damage'][1]['max']
                if tmp_attack.get('roles', None):
                    tmp['roles'] = list(tmp_attack.get('roles', None))
                    if tmp['roles'][0]['role'] == "role":
                        tmp['roles'][0]['role'] = "Variation-role 1"
                if tmp_attack.get('tags', None):
                    tmp['tags'] = list(tmp_attack.get('tags', None))
                    if tmp['tags'][0]['tag'] == "tag":
                        tmp['tags'][0]['tag'] = "Variation-tag 1"
                if tmp_attack.get('notes', None):
                    tmp['notes'] = [{'note': 'Variation-note BIG'}]
                    if tmp_attack['notes'][0]['note'] == "Variation-note BIG":
                        tmp['notes'][0]['note'] = "Variation-note 1"
                tmp_attack['variation'] = [tmp]

                # add new attack
                tmp2 = {
                    'title': {
                        'en': translateAttack(attack_id, lang="en"),
                        'de': translateAttack(attack_id, lang="de"),
                        'fr': translateAttack(attack_id, lang="fr"),
                        'ja': translateAttack(attack_id, lang="ja"),
                        'cn': translateAttack(attack_id, lang="cn"),
                        'ko': translateAttack(attack_id, lang="ko")
                    },
                    'title_id': attack_id,
                    'damage_type': attack['damage_type'],
                    'single_or_aoe': attack['type_id']
                }
                if tmp_attack.get('damage', None):
                    tmp2["damage"] = {}
                    try:
                        tmp2["damage"]["min"] = tmp_attack['damage']['min']
                        tmp2["damage"]["max"] = tmp_attack['damage']['max']
                    except:
                        tmp2["damage"]["min"] = tmp_attack['damage'][0]['min']
                        tmp2["damage"]["max"] = tmp_attack['damage'][1]['max']
                if tmp_attack.get('roles', None):
                    tmp2['roles'] = list(tmp_attack.get('roles', None))
                    if tmp2['roles'][0]['role'] == "role":
                        tmp2['roles'][0]['role'] = "Variation-role 1"
                if tmp_attack.get('tags', None):
                    tmp2['tags'] = list(tmp_attack.get('tags', None))
                    if tmp2['tags'][0]['tag'] == "tag":
                        tmp2['tags'][0]['tag'] = "Variation-tag 1"
                if tmp_attack.get('notes', None):
                    tmp2['notes'] = [{'note': 'Variation-note BIG'}]
                    if tmp_attack['notes'][0]['note'] == "Variation-note BIG":
                        tmp2['notes'][0]['note'] = "Variation-note 1"
                tmp_attack['variation'].append(tmp2)

                # delete old invalid entries
                tmp_attack = delete_invalid_entries(tmp_attack)
                existing_attacks[attack['name']] = 'variation'
            else:
                print_color_blue("\t\tNeed to add to variation: " + attack['name'], disable_blue_print)
                # find attack
                attack_index = None
                for i, old_attack in enumerate(old_enemy_data['attacks']):
                    if old_attack.get('title', {}).get('de', None) == attack['name']:
                        attack_index = i
                        break
                tmp_attack = old_enemy_data['attacks'][attack_index]
                # add new attack
                tmp = {
                    'title': {
                        'en': translateAttack(attack_id, lang="en"),
                        'de': translateAttack(attack_id, lang="de"),
                        'fr': translateAttack(attack_id, lang="fr"),
                        'ja': translateAttack(attack_id, lang="ja"),
                        'cn': translateAttack(attack_id, lang="cn"),
                        'ko': translateAttack(attack_id, lang="ko")
                    },
                    'title_id': attack_id,
                    'damage_type': attack['damage_type'],
                    'single_or_aoe': attack['type_id']
                }
                if tmp_attack['variation'][0].get('damage', None) or tmp_attack.get('damage', None):
                    tmp["damage"] = {}
                    try:
                        tmp["damage"]["min"] = tmp_attack['variation'][0].get('damage', {}).get('min', None) or tmp_attack.get('damage', {}).get('min', None)
                        tmp["damage"]["max"] = tmp_attack['variation'][0].get('damage', {}).get('max', None) or tmp_attack.get('damage', {}).get('max', None)
                    except:
                        tmp["damage"]["min"] = tmp_attack['variation'][0].get('damage', {})[0].get('min', None) or tmp_attack.get('damage', {})[0].get('min', None)
                        tmp["damage"]["max"] = tmp_attack['variation'][0].get('damage', {})[1].get('max', None) or tmp_attack.get('damage', {})[1].get('max', None)
                if tmp_attack['variation'][0].get('roles', None) or tmp_attack.get('roles', None):
                    tmp['roles'] = tmp_attack['variation'][0].get('roles', None) or tmp_attack.get('roles', None)
                if tmp_attack['variation'][0].get('tags', None) or tmp_attack.get('tags', None):
                    tmp['tags'] = tmp_attack['variation'][0].get('tags', None) or tmp_attack.get('tags', None)
                if tmp_attack['variation'][0].get('notes', None) or tmp_attack.get('notes', None):
                    tmp['notes'] = tmp_attack['variation'][0].get('notes', None) or tmp_attack.get('notes', None)
                tmp_attack['variation'].append(tmp)
        else:
            print_color_blue("\t\tAdded new entry for regular: " + attack['name'], disable_blue_print)
            # create new regular
            tmp = {
                'sortname': translateAttack(attack_id, lang="de"),
                'title': {
                    'en': translateAttack(attack_id, lang="en"),
                    'de': translateAttack(attack_id, lang="de"),
                    'fr': translateAttack(attack_id, lang="fr"),
                    'ja': translateAttack(attack_id, lang="ja"),
                    'cn': translateAttack(attack_id, lang="cn"),
                    'ko': translateAttack(attack_id, lang="ko")
                },
                'title_id': attack_id,
                'attack_in_use': 'false',
                'disable': 'false',
                'type': 'regular',
                'damage_type': attack['damage_type'],
                'phases': [{'phase': '09'}],
                'roles': [{'role': 'role'}],
                'tags': [{'tag': 'tag'}],
                'notes': [{'note': 'note'}],
                'single_or_aoe': attack['type_id']
            }
            if attack.get('damage', None):
                tmp["damage"] = {}
                tmp["damage"]["min"] = attack['damage']['min']
                tmp["damage"]["max"] = attack['damage']['max']

            if tmp['title']['de'] == "Attacke":
                tmp['attack_in_use'] = 'true'
                tmp['disable'] = 'true'
                tmp['roles'][0]['role'] = 'Tank'
                tmp['tags'][0]['tag'] = 'Auto-Angriff'
                del tmp['notes']

            if tmp['title']['de'].startswith("Unknown_"):
                tmp['disable'] = 'true'

            if enemy_type == "adds" and tmp.get("notes", None):
                del tmp['notes']
            if old_enemy_data.get('attacks', []) == []:
                old_enemy_data['attacks'] = []
            old_enemy_data['attacks'].append(tmp)

            existing_attacks[attack['name']] = 'regular'

    old_enemy_data = sort_list_of_skills(old_enemy_data)
    return old_enemy_data


def compare_status_ids(old_enemy_data, new_enemy_data, existing_debuffs, remove_debuff):
    for debuff_id in new_enemy_data.get('status', {}):
        for debuff in old_enemy_data.get('debuffs', {}):
            if type(debuff['title']) == str:
                debuff['title'] = addLanguageElements("status", debuff.get('title_id', None), debuff['title'])
            existing_debuffs[debuff['title']['de']] = ""
            if debuff_id == debuff.get('title_id', None):
                remove_debuff.append(debuff_id)
    return existing_debuffs, remove_debuff


def remove_status_from_list_if_found(remove_debuff, new_enemy_data):
    for x in remove_debuff:
        try:
            del new_enemy_data["status"][x]
        except Exception:
            pass
    return new_enemy_data


def merge_debuffs(old_enemy_data, new_enemy_data, enemy_type, saved_used_skills_to_ignore_in_last):
    remove_attack = []
    existing_debuffs = {}
    ignore_debuffs = ['130', '13d']
    obviouse_debuffs = ['82b', '82a', '82c', '9d4', '9d7', '95', '12', '38e', '828', '0e', '113', '06', '07', '01', '11', '282', 'ca', '11', '196', '1b6', '23c', '']
    # filter duplicate debuffs
    tmp_debuffs = []
    tmp_debuff_ids = []
    disable_yellow_print = False
    for x in old_enemy_data.get("debuffs", []):
        if x['title_id'] not in tmp_debuff_ids and x['title_id'] not in ignore_debuffs:
            saved_used_skills_to_ignore_in_last.append(x['title_id'])
            tmp_debuff_ids.append(x['title_id'])
            tmp_debuffs.append(x)
        else:
            print_color_yellow(f"Ignore Duplicate Debuff {x['title']}")
    disable_yellow_print = True
    old_enemy_data["debuffs"] = tmp_debuffs

    if not new_enemy_data.get('title', {}).get('de', "") == 'Unbekannte Herkunft':
        # remove debuffs already manually adjusted
        for debuff in saved_used_skills_to_ignore_in_last:
            try:
                del new_enemy_data.get("status", {})[debuff]
            except Exception:
                pass
        # comapre all skill ids
        existing_debuffs, remove_attack = compare_status_ids(old_enemy_data, new_enemy_data, existing_debuffs, remove_attack)
        # remove skill ids if they were found before
        new_enemy_data = remove_status_from_list_if_found(remove_attack, new_enemy_data)
    # comapre all skill ids
    existing_debuffs, remove_attack = compare_status_ids(old_enemy_data, new_enemy_data, existing_debuffs, remove_attack)
    # remove skill ids if they were found before
    new_enemy_data = remove_status_from_list_if_found(remove_attack, new_enemy_data)

    old_enemy_data = addSortName(old_enemy_data, "debuffs")
    # merge new keys
    if not old_enemy_data.get('debuffs', None):
        old_enemy_data['debuffs'] = []
    else:
        old_enemy_data['debuffs'] = sorted(old_enemy_data['debuffs'], key=itemgetter('sortname'))

    if not new_enemy_data.get('status', None):
        return old_enemy_data, saved_used_skills_to_ignore_in_last

    # get only status to make sorting easier
    new_enemy_data_status = new_enemy_data.get('status', {})

    # sort after the name key in subdicts
    new_enemy_data_status = OrderedDict(sorted(new_enemy_data_status.items(), key=lambda x: x[1]['name']))

    for status_id, status_data in new_enemy_data_status.items():
        if status_data['name'] == "":
            continue
        if status_id in obviouse_debuffs:
            tmp_status = addknowndebuff(status_id, status_data)
        else:
            tmp_status = {
                'sortname': translateAttack(status_id, _type=status, lang="de"),
                'title': {
                    'en': translateAttack(status_id, _type=status, lang="en"),
                    'de': translateAttack(status_id, _type=status, lang="de"),
                    'fr': translateAttack(status_id, _type=status, lang="fr"),
                    'ja': translateAttack(status_id, _type=status, lang="ja"),
                    'cn': translateAttack(status_id, _type=status, lang="cn"),
                    'ko': translateAttack(status_id, _type=status, lang="ko")
                },
                'title_id': status_id,
                'icon': status_data['icon'],
                'description': get_fixed_status_description(status_id),
                'debuff_in_use': 'false',
                'disable': 'false',
                #'damage_type': status_data['damage_type'],
                'phases': [{'phase': '09'}],
                'roles': [{'role': 'role'}],
                'tags': [{'tag': 'tag'}],
                'notes': [{'note': 'note'}]
            }
            if status_data.get('duration', None):
                tmp_status["durations"] = status_data['duration']
        old_enemy_data['debuffs'].append(tmp_status)
    old_enemy_data['debuffs'] = sorted(old_enemy_data['debuffs'], key=itemgetter('sortname'))
    return old_enemy_data, saved_used_skills_to_ignore_in_last


def addknowndebuff(status_id, status_data):
    tmp_status = {
        'sortname': translateAttack(status_id, _type=status, lang="de"),
        'title': {
            'en': translateAttack(status_id, _type=status, lang="en"),
            'de': translateAttack(status_id, _type=status, lang="de"),
            'fr': translateAttack(status_id, _type=status, lang="fr"),
            'ja': translateAttack(status_id, _type=status, lang="ja"),
            'cn': translateAttack(status_id, _type=status, lang="cn"),
            'ko': translateAttack(status_id, _type=status, lang="ko")
        },
        'title_id': status_id,
        'icon': status_data['icon'],
        'description': get_fixed_status_description(status_id),
        'debuff_in_use': 'true',
        'disable': 'false',
        #'damage_type': status_data['damage_type'],
        'phases': [{'phase': '09'}],
        'roles': [{'role': 'Alle'}],
        'tags': [{'tag': 'Common'}],
    }
    return tmp_status


def reorder_old_enemies(entry, enemy_type, old_enemies):
    # sort old enemies
    final_old_enemy = []

    # do this only for bosses as they will be reorderd in the same order as in the EXCEL file
    if entry.get(enemy_type, None) and enemy_type == "bosse":
        for enemy in entry[enemy_type]:
            for old_enemy_data in old_enemies:
                # if old title is empty set it to default
                if not old_enemy_data['title']:
                    old_enemy_data['title'] = UNKNOWNTITLE
                if type(old_enemy_data['title']) == str:
                    old_enemy_data['title'] = addLanguageElements("bnpc", old_enemy_data['enemy_id'], old_enemy_data['title'])
                if old_enemy_data['title'].get('de', "").lower() == enemy.lower():
                    final_old_enemy.append(old_enemy_data)

    # if you created a new order, apply it
    if not final_old_enemy == []:
        old_enemies = final_old_enemy
    return old_enemies


def addLanguageElements(_type, _id, ori_value):
    if _type == "bnpc":
        data = bnpcname
        field = "Singular"
    if _type == "action":
        data = action
        field = "Name"
        _id = str(int(_id, 16))
    if _type == "status":
        data = status
        field = "Name"
        _id = str(int(_id, 16))

    if _id:
        return {
            'en': fixCaptilaziationAndRomanNumerals(data[_id][f'{field}_en']),
            'de': fixCaptilaziationAndRomanNumerals(ori_value),
            'fr': fixCaptilaziationAndRomanNumerals(data[_id][f'{field}_fr']),
            'ja': fixCaptilaziationAndRomanNumerals(data[_id][f'{field}_ja']),
            'cn': fixCaptilaziationAndRomanNumerals(data[_id][f'{field}_cn']),
            'ko': fixCaptilaziationAndRomanNumerals(data[_id][f'{field}_ko']),
        }
    return UNKNOWNTITLE


def workOnOldEnemies(guide_data, entry, enemy_type, old_enemies, logdata_instance_content):
    empty_enemy_available = False
    if old_enemies:
        print_color_red(f"\tStart looking at old_enem {enemy_type}", disable_red_print)
        old_enemies = reorder_old_enemies(entry, enemy_type, old_enemies)

        saved_used_skills_to_ignore_in_last = []
        for i, old_enemy_data in enumerate(old_enemies):
            if type(old_enemy_data['title']) == str:
                old_enemy_data['title'] = addLanguageElements("bnpc", old_enemy_data['enemy_id'], old_enemy_data['title'])
            if old_enemy_data['title']['de'] == "":
                continue
            elif old_enemy_data['title']['de'] == "Unbekannte Herkunft":
                empty_enemy_available = True

            old_enemy_data['title']['de'] = fixCaptilaziationAndRomanNumerals(old_enemy_data['title']['de'])
            print_color_yellow(f"\tWork on '{old_enemy_data['title']['de']}'", disable_yellow_print)
            empty_enemy_data = {}
            try:
                new_enemy_data = logdata_instance_content[old_enemy_data['title']['de']]
                if enemy_type == 'bosse' and len(old_enemies) == i + 1:
                    empty_enemy_data = logdata_instance_content['']
            except Exception as e:
                # print(logdata_instance_content)
                if not old_enemy_data['title']['de'] == "Unbekannte Herkunft":
                    print(f"Could not find {old_enemy_data['title']['de']} in logdata")
                new_enemy_data = {}
            new_enemy_data_c = copy.deepcopy({**{}, **new_enemy_data})

            # get enemy attacks and debuffs for all enemies with a name
            old_enemy_data["enemy_id"] = new_enemy_data.get("id", old_enemy_data.get('enemy_id', ""))

            if new_enemy_data.get("minHP", None) or new_enemy_data.get("maxHP", None):
                old_enemy_data["hp"] = {}
                old_enemy_data["hp"]["min"] = new_enemy_data.get("minHP", "")
                old_enemy_data["hp"]["max"] = new_enemy_data.get("maxHP", "")

            old_enemy_data = merge_attacks(old_enemy_data, new_enemy_data, enemy_type)
            old_enemy_data, saved_used_skills_to_ignore_in_last = merge_debuffs(old_enemy_data, new_enemy_data, enemy_type, saved_used_skills_to_ignore_in_last)

            guide_data += add_Enemy(old_enemy_data, enemy_type, new_enemy_data_c)

            # this try catch will remove enemy from logdata, so it wont be readded later on
            try:
                if old_enemy_data['title']['de'].title() == "Hesperos I":
                    del logdata_instance_content["Hesperos"]
            except:
                print_color_red("!Could not remove enemy: Hesperos", disable_red_print)

            try:
                del logdata_instance_content[fixCaptilaziationAndRomanNumerals(old_enemy_data['title']['de'])]
            except Exception as e:
                print_color_red("Could not remove enemy: " + old_enemy_data['title']['de'], disable_red_print)

            # handle enemy if name is ""
            if not empty_enemy_data == {}:
                empty_enemy_available = True
                base_data = {'title': UNKNOWNTITLE, 'id': f"boss0{i+2}", "sequence": [{"phase": "09"}], 'debuffs': []}
                old_enemy_data = {'title': UNKNOWNTITLE, 'id': f"boss0{i+2}", "sequence": [{"phase": "09"}], 'debuffs': []}
                old_enemy_data = merge_attacks(old_enemy_data, empty_enemy_data, enemy_type)
                old_enemy_data, saved_used_skills_to_ignore_in_last = merge_debuffs(old_enemy_data, empty_enemy_data, enemy_type, saved_used_skills_to_ignore_in_last)

                if not old_enemy_data == base_data:
                    guide_data += add_Enemy(old_enemy_data, enemy_type, new_enemy_data_c)
    return guide_data, empty_enemy_available


def workOnLogDataEnemies(entry, enemy_type, logdata_instance_content, empty_enemy_available):
    guide_data = ""
    if logdata_instance_content != {}:
        print_color_red(f"\t Start looking at logdata {enemy_type}", disable_red_print)
        counter = 0
        if logdata_instance_content is None:
            return guide_data

        # the array is used to sort bosses
        enemies_to_add = []
        empty_enemy = None
        for i, enemy in enumerate(logdata_instance_content, 1):
            if enemy == "" and enemy_type == 'adds':
                continue
            print_color_yellow(f"\tWork on '{enemy}'", disable_yellow_print)
            # for bosses, only write bosses, for adds skip bosse
            lower_enemy_list = [x.lower() for x in entry[enemy_type]]
            lower_boss_list = [x.lower() for x in entry['bosse']]
            lower_enemy_list.append("")
            if (enemy_type == 'bosse' and enemy.lower() not in lower_enemy_list) or (enemy_type == 'adds' and enemy.lower() in lower_boss_list):
                logger.debug(f"Skip {enemy} ({enemy_type}) -> " + str(enemy.lower() not in lower_enemy_list))
                continue
            counter = counter + 1
            new_enemy_data = logdata_instance_content.get(enemy, "Unknown_")
            if new_enemy_data == {}:
                new_enemy_data = {'skill': {}}
            new_enemy_data_c = copy.deepcopy({**{}, **new_enemy_data})
            # create new template file to merge against
            enemy_id = new_enemy_data.get("id", "")
            de_and_sort_name = enemy if enemy != "" else "Unbekannte Herkunft"
            tmp = {
                "sortname": de_and_sort_name,
                "title": {
                    "en": getBnpcName(enemy, enemy_id, "en") if enemy != "" else "Unknown Source",
                    "de": de_and_sort_name,
                    "fr": getBnpcName(enemy, enemy_id, "fr") if enemy != "" else "Unknown Source",
                    "ja": getBnpcName(enemy, enemy_id, "ja") if enemy != "" else "Unknown Source",
                    "cn": getBnpcName(enemy, enemy_id, "cn") if enemy != "" else "Unknown Source",
                    "ko": getBnpcName(enemy, enemy_id, "ko") if enemy != "" else "Unknown Source"
                },
                "enemy_id": enemy_id,
                "id": f"{enemy_type[:-1]}{counter:02d}",
                "attacks": [],
                "debuffs": []
            }

            if new_enemy_data.get("minHP", None) or new_enemy_data.get("maxHP", None):
                tmp["hp"] = {}
                tmp["hp"]["min"] = new_enemy_data.get("minHP", "")
                tmp["hp"]["max"] = new_enemy_data.get("maxHP", "")

            enemy_data = merge_attacks(tmp, new_enemy_data, enemy_type)
            enemy_data['attacks'] = sorted(enemy_data['attacks'], key=itemgetter('sortname'))

            enemy_data, saved_used_skills_to_ignore_in_last = merge_debuffs(tmp, new_enemy_data, enemy_type, [])
            enemy_data['debuffs'] = sorted(enemy_data['debuffs'], key=itemgetter('sortname'))
            if enemy == "":
                empty_enemy = enemy_data
            else:
                tmp_guide_data_from_enemy = add_Enemy(enemy_data, enemy_type, new_enemy_data_c)
                enemies_to_add.append(tmp_guide_data_from_enemy)

        # this will add bosses in order as specified and add adds in order they were added to the array
        if enemy_type == 'bosse':
            for boss in entry['bosse']:
                for e in enemies_to_add:
                    if f'      de: "{boss}"' in e:
                        guide_data += e
                        break
        else:
            for e in enemies_to_add:
                guide_data += e

        if empty_enemy and not empty_enemy_available:
            if empty_enemy.get("attacks", None) or empty_enemy.get("debuffs", None):
                guide_data += add_Enemy(empty_enemy, "bosse", new_enemy_data_c)
    return guide_data


def ugly_fix_enemy_data(enemy_data, new_enemy_data):
    for attack in enemy_data.get('attacks', []):
        if attack['type'] in ["variation", "combo"]:
            for action in attack[attack['type']]:
                if new_enemy_data.get("skill", {}).get(action["title_id"], None):
                    x = new_enemy_data.get("skill", {}).get(action["title_id"], {})
                    if x.get("damage", {}).get("min", None) or x.get("damage", {}).get("max", None):
                        action['damage'] = {}
                        action['damage']['min'] = x.get("damage", {}).get("min", None)
                        action['damage']['max'] = x.get("damage", {}).get("max", None)
        elif attack['type'] == "regular":
            if new_enemy_data.get("skill", {}).get(attack["title_id"], None):
                x = new_enemy_data.get("skill", {}).get(attack["title_id"], {})
                if x.get("damage", {}).get("min", None) or x.get("damage", {}).get("max", None):
                    attack['damage'] = {}
                    attack['damage']['min'] = x.get("damage", {}).get("min", None)
                    attack['damage']['max'] = x.get("damage", {}).get("max", None)
    return enemy_data


def setMultipleLanguageStrings(guide_data, elementName, elemntArray, spaces):
    for lang in LANGUAGES:
        if type(elemntArray[elementName]) == str:
            continue
        if elemntArray[elementName].get(f"{lang}", None):
            guide_data += f'{spaces}{lang}: "{elemntArray[elementName][lang]}"\n'
    return guide_data


def add_Debuff(guide_data, debuff, enemy_type):
    guide_data += f'      - title:\n'
    guide_data = setMultipleLanguageStrings(guide_data, "title", debuff, "          ")
    guide_data += f'        title_id: "{debuff["title_id"]}"\n'
    guide_data += f'        icon: "{getImage(debuff["icon"])}"\n'
    guide_data += f'        description: "{debuff["description"]}"\n'
    if debuff.get("durations", None):
        guide_data += f'        durations: {debuff["durations"]}\n'
    guide_data += f'        debuff_in_use: "{debuff["debuff_in_use"] or "false"}"\n'
    guide_data += f'        disable: "{debuff["disable"] or "false"}"\n'
    #guide_data += f'        damage_type: "{debuff["damage_type"] or "None"}"\n'

    if debuff.get('phases', None):
        guide_data += f'        phases:\n'
        for phase in debuff.get('phases', {}):
            guide_data += f'          - phase: "{int(phase["phase"]):02d}"\n'

    if debuff["title"]['de'].startswith("Unknown_"):
        return guide_data

    if debuff.get('roles', None):
        guide_data += f'        roles:\n'
        for role in debuff.get('roles', {}):
            guide_data += f'          - role: "{role["role"]}"\n'

    if debuff.get('tags', None):
        guide_data += f'        tags:\n'
        for tag in debuff.get('tags', {}):
            guide_data += f'          - tag: "{tag["tag"]}"\n'

    if debuff.get('notes', None):
        guide_data += f'        notes:\n'
        for note in debuff.get('notes', {}):
            guide_data += f'          - note: "{note["note"]}"\n'

    if debuff.get('videos', None):
        guide_data += f'        videos:\n'
        for video in debuff.get('videos', {}):
            guide_data += f'          - url: "{video["url"]}"\n'
    return guide_data


def add_regular_Attack(guide_data, attack, enemy_type):
    guide_data += f'      - title:\n'
    guide_data = setMultipleLanguageStrings(guide_data, "title", attack, "          ")
    guide_data += f'        title_id: "{attack["title_id"]}"\n'
    guide_data += f'        attack_in_use: "{attack["attack_in_use"] or "false"}"\n'
    guide_data += f'        disable: "{attack.get("disable", "false")}"\n'
    guide_data += f'        type: "{attack["type"] or "regular"}"\n'
    guide_data += f'        damage_type: "{attack.get("damage_type", None)}"\n'

    if attack.get('damage', None):
        guide_data += f'        damage:\n'
        try:
            guide_data += f'          - min: {attack["damage"]["min"] or "None"}\n'
            guide_data += f'          - max: {attack["damage"]["max"] or "None"}\n'
        except TypeError:
            guide_data += f'          - min: {attack["damage"][0]["min"] or "None"}\n'
            guide_data += f'          - max: {attack["damage"][1]["max"] or "None"}\n'

    if attack.get('phases', None):
        guide_data += f'        phases:\n'
        for phase in attack.get('phases', {}):
            guide_data += f'          - phase: "{int(phase["phase"]):02d}"\n'

    if attack["title"]['de'].startswith("Unknown_"):
        return guide_data

    if attack.get('roles', None):
        guide_data += f'        roles:\n'
        for role in attack.get('roles', {}):
            guide_data += f'          - role: "{role["role"]}"\n'

    if attack.get('tags', None):
        guide_data += f'        tags:\n'
        for tag in attack.get('tags', {}):
            guide_data += f'          - tag: "{tag["tag"]}"\n'
        if attack.get('single_or_aoe', None):
            x = "AoE" if attack["single_or_aoe"] == "22" else "Single"
            guide_data += f'          - tag: "{x}"\n'

    if attack.get('notes', None):
        guide_data += f'        notes:\n'
        for note in attack.get('notes', {}):
            guide_data += f'          - note: "{note["note"]}"\n'

    if attack.get('images', None):
        guide_data += f'        images:\n'
        for image in attack.get('images', {}):
            guide_data += f'          - url: "{image["url"]}"\n'
            guide_data += f'            alt: "{image.get("alt", None) or image["url"]}"\n'
            guide_data += f'            height: "{image.get("height", None) or "350px"}"\n'

    if attack.get('videos', None):
        guide_data += f'        videos:\n'
        for video in attack.get('videos', {}):
            guide_data += f'          - url: "{video["url"]}"\n'
    return guide_data


def add_variation_Attack(guide_data, attack, enemy_type):
    guide_data += f'      - title:\n'
    guide_data = setMultipleLanguageStrings(guide_data, "title", attack, "          ")
    guide_data += f'        attack_in_use: "{attack["attack_in_use"] or "false"}"\n'
    guide_data += f'        disable: "{attack.get("disable", "false")}"\n'
    guide_data += f'        type: "{attack["type"] or "regular"}"\n'

    if attack.get('phases', None):
        guide_data += f'        phases:\n'
        for phase in attack.get('phases', {}):
            guide_data += f'          - phase: "{int(phase["phase"]):02d}"\n'

    if (attack.get('notes', None) and enemy_type != "adds") or (attack.get('notes', None) and enemy_type == "adds" and attack.get('notes', [{}])[0].get("note", None) != "Variation-note BIG"):
        guide_data += f'        notes:\n'
        for note in attack.get('notes', {}):
            guide_data += f'          - note: "{note["note"]}"\n'

    if attack.get('variation', None):
        guide_data += f'        variation:\n'
        for variation in attack.get('variation', {}):
            #guide_data += f'          - title:\n'
            #guide_data = setMultipleLanguageStrings(guide_data, "title", variation, "              ")
            guide_data += f'          - title_id: "{variation["title_id"]}"\n'
            guide_data += f'            damage_type: "{variation.get("damage_type", None)}"\n'

            # print_color_yellow(variation)
            if variation.get('damage', None):
                guide_data += f'            damage:\n'
                try:
                    guide_data += f'              - min: {variation["damage"]["min"] or "None"}\n'
                    guide_data += f'              - max: {variation["damage"]["max"] or "None"}\n'
                except TypeError:
                    guide_data += f'              - min: {variation["damage"][0]["min"] or "None"}\n'
                    guide_data += f'              - max: {variation["damage"][1]["max"] or "None"}\n'

            if variation.get('roles', None):
                guide_data += f'            roles:\n'
                for role in variation.get('roles', {}):
                    guide_data += f'              - role: "{role["role"]}"\n'

            if variation.get('tags', None):
                guide_data += f'            tags:\n'
                for tag in variation.get('tags', {}):
                    guide_data += f'              - tag: "{tag["tag"]}"\n'
                if variation.get('single_or_aoe', None):
                    x = "AoE" if variation["single_or_aoe"] == "22" else "Single"
                    guide_data += f'              - tag: "{x}"\n'

            if variation.get('notes', None) and enemy_type != "adds":
                guide_data += f'            notes:\n'
                for note in variation.get('notes', {}):
                    guide_data += f'              - note: "{note["note"]}"\n'

    if attack.get('images', None):
        guide_data += f'        images:\n'
        for image in attack.get('images', {}):
            guide_data += f'          - url: "{image["url"]}"\n'
            guide_data += f'            alt: "{image.get("alt", None) or image["url"]}"\n'
            guide_data += f'            height: "{image.get("height", None) or "350px"}"\n'

    if attack.get('videos', None):
        guide_data += f'        videos:\n'
        for video in attack.get('videos', {}):
            guide_data += f'          - url: "{video["url"]}"\n'

    return guide_data


def add_combo_Attack(guide_data, attack, enemy_type):
    guide_data += f'      - title:\n'
    guide_data = setMultipleLanguageStrings(guide_data, "title", attack, "          ")
    guide_data += f'        attack_in_use: "{attack["attack_in_use"] or "false"}"\n'
    guide_data += f'        disable: "{attack["disable"] or "false"}"\n'
    guide_data += f'        type: "{attack["type"] or "regular"}"\n'

    if attack.get('phases', None):
        guide_data += f'        phases:\n'
        for phase in attack.get('phases', {}):
            guide_data += f'          - phase: "{int(phase["phase"]):02d}"\n'

    if (attack.get('notes', None) and enemy_type != "adds") or (attack.get('notes', None) and enemy_type == "adds" and attack.get('notes', [{}])[0].get("note", None) != "Variation-note BIG"):
        guide_data += f'        notes:\n'
        for note in attack.get('notes', {}):
            guide_data += f'          - note: "{note["note"]}"\n'

    if attack.get('combo', None):
        guide_data += f'        combo:\n'
        for combo in attack.get('combo', {}):
            guide_data += f'          - title: "{combo["title"]}"\n'
            guide_data += f'            title_id: "{combo["title_id"]}"\n'
            guide_data += f'            damage_type: "{combo["damage_type"]}"\n'

            if combo.get('damage', None):
                guide_data += f'            damage:\n'
                guide_data += f'              - min: {combo["damage"]["min"] or "None"}\n'
                guide_data += f'              - max: {combo["damage"]["max"] or "None"}\n'

            if combo.get('roles', None):
                guide_data += f'            roles:\n'
                for role in combo.get('roles', {}):
                    guide_data += f'              - role: "{role["role"]}"\n'

            if combo.get('tags', None):
                guide_data += f'            tags:\n'
                for tag in combo.get('tags', {}):
                    guide_data += f'              - tag: "{tag["tag"]}"\n'
                if combo.get('single_or_aoe', None):
                    x = "AoE" if combo["single_or_aoe"] == "22" else "Single"
                    guide_data += f'              - tag: "{x}"\n'

            if combo.get('notes', None) and enemy_type != "adds":
                guide_data += f'            notes:\n'
                for note in combo.get('notes', {}):
                    guide_data += f'              - note: "{note["note"]}"\n'

    if attack.get('images', None):
        guide_data += f'        images:\n'
        for image in attack.get('images', {}):
            guide_data += f'          - url: "{image["url"]}"\n'
            guide_data += f'            alt: "{image.get("alt", None) or image["url"]}"\n'
            guide_data += f'            height: "{image.get("height", None) or "350px"}"\n'

    if attack.get('videos', None):
        guide_data += f'        videos:\n'
        for video in attack.get('videos', {}):
            guide_data += f'          - url: "{video["url"]}"\n'

    return guide_data


def add_Sequence(guide_data, data):
    guide_data += "    sequence:\n"
    for phase in data['sequence']:
        guide_data += f"      - phase: \"{phase['phase']}\"\n"
        if phase.get('name', None):
            guide_data += f"        name: \"{phase['name']}\"\n"

        if phase.get('alerts', None):
            guide_data += "        alerts:\n"
            for alert in phase['alerts']:
                guide_data += f"          - alert: \"{alert['alert']}\"\n"

        if phase.get('toolbox', None):
            guide_data += "        toolbox:\n"
            for toolbox in phase['toolbox']:
                guide_data += f"          - link: \"{toolbox['link']}\"\n"
                guide_data += f"            name: \"{toolbox['name']}\"\n"

        if phase.get('mechanics', None):
            guide_data += "        mechanics:\n"
            for mechanic in phase['mechanics']:
                guide_data += f"          - title: \"{mechanic['title']}\"\n"
                if mechanic.get('notes', None):
                    guide_data += f"            notes:\n"
                    for note in mechanic['notes']:
                        guide_data += f"              - note: \"{note['note']}\"\n"

        if phase.get('attacks', None):
            guide_data += "        attacks:\n"
            for attack in phase['attacks']:
                guide_data += f"          - attack: \"{attack['attack']}\"\n"

        if phase.get('images', None):
            guide_data += "        images:\n"
            for image in phase['images']:
                guide_data += f"          - url: \"{image['url']}\"\n"
                guide_data += f"            alt: \"{image.get('alt', image['url'])}\"\n"
                guide_data += f"            height: \"{image.get('height', '250px')}\"\n"

        if phase.get('videos', None):
            guide_data += "        videos:\n"
            for video in phase['videos']:
                guide_data += f"          - url: \"{video['url']}\"\n"
    return guide_data


def add_Enemy(enemy_data, enemy_type, new_enemy_data):
    guide_data = ""
    enemy_data = ugly_fix_enemy_data(enemy_data, new_enemy_data)
    guide_data += f'  - title:\n'
    guide_data = setMultipleLanguageStrings(guide_data, "title", enemy_data, "      ")
    if type(enemy_data.get("enemy_id", "")) == list:
        guide_data += f'    enemy_id: "{", ".join(enemy_data.get("enemy_id", ""))}"\n'
    else:
        guide_data += f'    enemy_id: "{enemy_data.get("enemy_id", "")}"\n'
    guide_data += f'    id: "{enemy_data["id"]}"\n'

    if enemy_data.get('hp', None):
        guide_data += f'    hp:\n'
        try:
            guide_data += f'      - min: {enemy_data["hp"]["min"] or "None"}\n'
            guide_data += f'      - max: {enemy_data["hp"]["max"] or "None"}\n'
        except TypeError:
            guide_data += f'      - min: {enemy_data["hp"][0]["min"] or "None"}\n'
            guide_data += f'      - max: {enemy_data["hp"][1]["max"] or "None"}\n'
    if enemy_data.get("attacks", None):
        guide_data += '    attacks:\n'
        for attack in enemy_data["attacks"]:
            if attack["type"] == "regular":
                guide_data = add_regular_Attack(guide_data, attack, enemy_type)
            elif attack["type"] == "variation":
                guide_data = add_variation_Attack(guide_data, attack, enemy_type)
            elif attack["type"] == "combo":
                guide_data = add_combo_Attack(guide_data, attack, enemy_type)
    if enemy_data.get("debuffs", None):
        guide_data += '    debuffs:\n'
        for debuff in enemy_data["debuffs"]:
            guide_data = add_Debuff(guide_data, debuff, enemy_type)
    if enemy_data.get("sequence", None):
        guide_data = add_Sequence(guide_data, enemy_data)
    else:
        if enemy_type == "bosse":
            guide_data = add_Sequence(guide_data, example_sequence)
        else:
            guide_data = add_Sequence(guide_data, example_add_sequence)
    return guide_data


def check_Enemy(entry, enemy_type, logdata_instance_content, old_enemies):
    guide_data = ""
    if old_enemies == {} and logdata == {}:
        return guide_data
    guide_data += "bosses:\n" if enemy_type == "bosse" else "adds:\n"
    guide_data, empty_enemy_available = workOnOldEnemies(guide_data, entry, enemy_type, old_enemies, logdata_instance_content)
    print_color_yellow(guide_data, disable_yellow_print)
    guide_data += workOnLogDataEnemies(entry, enemy_type, logdata_instance_content, empty_enemy_available)
    print_color_blue(guide_data, disable_blue_print)
    return guide_data


# Notizen, Bosse und Adds
def addGuide(entry, old_data, logdata_instance_content):
    guide_data = ""
    # add mechanics
    guide_data += check_Mechanics(entry, old_data.get('mechanics', None))
    print_color_green(f"Work on '{entry['title_de']}'", disable_green_print)
    guide_data += check_Enemy(entry, "bosse", logdata_instance_content, old_data.get('bosses', {}))
    guide_data += check_Enemy(entry, "adds", logdata_instance_content, old_data.get('adds', {}))
    return guide_data


####################################################################################################################


def addHeader(entry, old_data, music, contentzoneid):
    header_data, entry = rewrite_content_even_if_exists(entry, old_data.get('wip', False))
    header_data, cmt = addContentZoneIdToHeader(header_data, contentzoneid, entry)
    header_data += addGroupCollections(cmt, entry)
    if music:
        header_data = addMusic(header_data, music)
    return header_data


def write_content_to_file(entry, filename, old_data):
    logdata_instance_content, music, contentzoneid = getDataFromLogfile(entry)
    filedata = '---\n'
    filedata += addHeader(entry, old_data, music, contentzoneid)
    filedata += addGuide(entry, old_data, logdata_instance_content)
    filedata += '---'
    filedata += '\n'
    writeFileIfNoDifferent(filename, filedata)


def run(sheet, max_row, max_column, elements, orderedContent):
    # for every row do:
    for i in range(2, max_row):
        try:
            # comment the 2 line out to filter fo a specific line, numbering starts with 1 like it is in excel
            #if i not in [0]:
            #    continue
            entry = getEntryData(sheet, max_column, i, elements, orderedContent)
            logger.info(pretty_json(entry))
            # if the done collumn is not prefilled
            if entry["exclude"] == "end":
                print("END FLAG WAS FOUND!")
                sys.exit(0)
            if not (entry["exclude"] or entry["done"]):
                logger.debug(pretty_json(entry))
                filename = f"{entry['categories']}_new/{entry['instanceType']}/{entry['date'].replace('.', '-')}--{entry['patchNumber']}--{entry['sortid'].zfill(5)}--{entry['slug'].replace(',', '')}.md"
                existing_filename = f"{entry['categories']}/{entry['instanceType']}/{entry['date'].replace('.', '-')}--{entry['patchNumber']}--{entry['sortid'].zfill(5)}--{entry['slug'].replace(',', '')}.md"
                old_data = get_old_content_if_file_is_found(existing_filename)
                # if old file was found, replace filename to save
                if not old_data == {}:
                    filename = existing_filename
                    # logger.info(pretty_json(old_data))
                try_to_create_file(filename)
                write_content_to_file(entry, filename, old_data)
        except Exception as e:
            logger.critical(f"Error when handeling '{filename}' with line id '{i}' ({e})")
            traceback.print_exception(*sys.exc_info())


if __name__ == "__main__":
    sheet, max_row, max_column = read_xlsx_file()
    # change into _posts dir
    os.chdir("./_posts")
    # first run to create all files
    orderedContent = getPrevAndNextContentOrder(sheet, XLSXELEMENTS, max_row)
    logger.debug(orderedContent)
    run(sheet, max_row, max_column, XLSXELEMENTS, orderedContent)
    # csgf needs also to run from posts dir
    csgf.run()
    # move back to DEVPOCKETGUIDE dir
    os.chdir("..")
    gl.run()
